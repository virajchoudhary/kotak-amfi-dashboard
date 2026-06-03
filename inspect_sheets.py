import sys
sys.path.append("backend")
import openpyxl

wb = openpyxl.load_workbook("AMFI_MOM DATA - Apr'25 to Mar26.xlsx")
ws_flat = wb["AMFI-Mar'25 to Mar'26"]
print("Flat sheet columns 108 to 147:")
for col in range(108, 148):
    if col <= ws_flat.max_column:
        print(f"Col {col}: Row 2='{ws_flat.cell(2, col).value}', Row 3='{ws_flat.cell(3, col).value}'")
