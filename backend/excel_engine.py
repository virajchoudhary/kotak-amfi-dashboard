import calendar
import csv
from collections import defaultdict
from copy import copy
from datetime import datetime
from functools import lru_cache
import io
import logging
import os
from pathlib import Path
import re
import sqlite3
import threading
import zipfile

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.properties import Outline
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.views import Selection

BACKEND_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = BACKEND_DIR.parent
TEMPLATE_PATH = BACKEND_DIR / "data" / "template file.xlsx"
SIP_HISTORY_PATH = Path(__file__).resolve().parent / "data" / "sip_history.csv"
_UPLOAD_LOCK = threading.Lock()
_DASHBOARD_CACHE: dict[str, tuple] = {}
_DASHBOARD_CACHE_LOCK = threading.Lock()
LOGGER = logging.getLogger("amfi_dashboard.excel_engine")
SHEET_SIP = "AMFI-SIP"
SHEET_FLAT = "AMFI-Mar'25 to Mar'26"
SHEET_FORM = "AMFI-Mar'25 to Jan'26-AMFI form"
SHEET_NS = "NS-Analysis"
SHEET_FLAT_PREFIX = "AMFI-Mar'25 to "
SHEET_FORM_PREFIX = "AMFI-Mar'25 to "
SHEET_FORM_SUFFIX = "-AMFI form"
HEADER_ROW = 3
MONTH_ROW = 2
DATA_START_ROW = 4
ATTRIBUTE_COLUMN_COUNT = 6
SUMMARY_HEADERS = ("Latest Net AUM", "Monthly Sales", "FYTD Sales")
CATEGORY_NAMES = (
    "Active Equity (incl Hybrid)",
    "Arb/ESF",
    "Active Debt (incl Hybrid)",
    "Cash",
    "Passive-E/D",
    "Interval Schemes",
)

NS_EQUITY_TEMPLATE_ROWS = (
    (10, 6),   # Flexi Cap Fund
    (11, 7),   # Mid Cap Fund
    (12, 9),   # Large & Mid Cap Fund
    (13, 8),   # Small Cap Fund
    (14, 11),  # Large Cap Fund
    (15, 12),  # Multi Cap Fund
    (16, 15),  # Focused Fund
    (17, 10),  # Sectoral/Thematic Funds
    (18, 16),  # Value Fund/Contra Fund
    (19, 17),  # Childrens Fund
    (20, 19),  # Retirement Fund
    (21, 18),  # Dividend Yield Fund
    (22, 20),  # Other Equity Schemes
    (23, 21),  # Other Schemes
    (24, 23),  # ELSS-Close Ended
    (25, 24),  # ELSS
)
NS_HYBRID_TEMPLATE_ROWS = (
    (30, 14),  # Balanced Hybrid Fund/Aggressive Hybrid Fund
    (31, 13),  # Dynamic Asset Allocation/Balanced Advantage Fund
    (32, 5),   # Multi Asset Allocation Fund
)

MONTHS = {
    "jan": 1, "january": 1, "feb": 2, "february": 2, "mar": 3, "march": 3,
    "apr": 4, "april": 4, "may": 5, "jun": 6, "june": 6, "jul": 7, "july": 7,
    "aug": 8, "august": 8, "sep": 9, "sept": 9, "september": 9, "oct": 10,
    "october": 10, "nov": 11, "november": 11, "dec": 12, "december": 12,
}
MONTH_ABBR = {v: k.title() for k, v in {"jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6, "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12}.items()}

METRIC_ORDER = [
    "no_schemes", "folios", "funds_mobilized", "redemption", "net_inflow",
    "net_aum", "avg_aum", "seg_portfolios", "seg_aum",
]

METRIC_LABELS = {
    "no_schemes": "No. of Schemes as on {date}",
    "folios": "No. of Folios as on {date}",
    "funds_mobilized": "Funds Mobilized for the month of {month} {year} (INR in crore)",
    "redemption": "Repurchase/Redemption for the month of {month} {year} (INR in crore)",
    "net_inflow": "Net Inflow (+ve)/Outflow (-ve) for the month of {month} {year} (INR in crore)",
    "net_aum": "Net Assets Under Management as on {date} (INR in crore)",
    "avg_aum": "Average Net Assets Under Management for the month {month} {year} (INR in crore)",
    "seg_portfolios": "No. of segregated portfolios created as on {date}",
    "seg_aum": "Net Assets Under Management in segregated portfolio as on {date} (INR in crore)",
}

HYBRID_SCHEMES = {
    "multi asset allocation fund",
    "dynamic asset allocation balanced advantage fund",
    "balanced hybrid fund aggressive hybrid fund",
    "conservative hybrid fund",
}

_FORM_SUBTOTAL_ROWS = {
    22: [(6, 21)],
    36: [(25, 35)],
    45: [(39, 44)],
    50: [(48, 49)],
    57: [(53, 56)],
    59: [(6, 21), (25, 35), (39, 44), (48, 49), (53, 56)],
    67: [(63, 66)],
    72: [(70, 71)],
    76: [(63, 66), (70, 71), (74, 74)],
    85: [(79, 83)],
    87: [(59, 59), (76, 76), (85, 85)],
}

def database_path() -> Path:
    configured = os.getenv("AMFI_DB_PATH")
    if configured:
        return Path(configured).expanduser().resolve()
    return BACKEND_DIR / "amfi.db"

def _env_bool(name: str, default: bool = False) -> bool:
    value = os.getenv(name)
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "on"}

DB_PATH = database_path()

def get_db_connection():
    db_path = database_path()
    if not db_path.parent.exists():
        raise RuntimeError(f"Database directory does not exist: {db_path.parent}")
    conn = sqlite3.connect(db_path, timeout=10.0)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    return conn

def template_bytes() -> bytes:
    return _sanitize_xlsx_package(TEMPLATE_PATH.read_bytes())

def _sanitize_xlsx_package(data: bytes) -> bytes:
    """Remove stale external-link package parts that make Excel prompt/repair."""
    source = io.BytesIO(data)
    target = io.BytesIO()
    external_link_re = re.compile(
        r'<Relationship\b[^>]*Type="[^"]*/externalLink"[^>]*/>',
        flags=re.IGNORECASE,
    )
    calc_chain_re = re.compile(
        r'<Relationship\b[^>]*Type="[^"]*/calcChain"[^>]*/>',
        flags=re.IGNORECASE,
    )
    external_reference_re = re.compile(
        r'<externalReferences\b[^>]*>.*?</externalReferences>',
        flags=re.IGNORECASE | re.DOTALL,
    )
    external_content_type_re = re.compile(
        r'<Override\b[^>]*PartName="/xl/externalLinks/[^"]*"[^>]*/>',
        flags=re.IGNORECASE,
    )
    calc_chain_content_type_re = re.compile(
        r'<Override\b[^>]*PartName="/xl/calcChain.xml"[^>]*/>',
        flags=re.IGNORECASE,
    )
    with zipfile.ZipFile(source, "r") as zin, zipfile.ZipFile(target, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            name = item.filename
            if name.startswith("xl/externalLinks/") or name == "xl/calcChain.xml":
                continue
            content = zin.read(name)
            if name.endswith((".xml", ".rels")):
                text = content.decode("utf-8", errors="ignore")
                text = text.replace("[1]Analytics!", "Analytics!")
                text = text.replace("Analytics!#REF!", "Analytics!$D$4")
                text = text.replace("<c:v>#REF!</c:v>", "<c:v>AUM</c:v>")
                if name == "xl/charts/chart3.xml":
                    text = text.replace("Analytics!$A$28:$A$40", "Analytics!$P$28:$P$40")
                    text = text.replace("Analytics!$B$28:$B$40", "Analytics!$Q$28:$Q$40")
                if name == "xl/workbook.xml":
                    text = external_reference_re.sub("", text)
                if name == "xl/_rels/workbook.xml.rels":
                    text = external_link_re.sub("", text)
                    text = calc_chain_re.sub("", text)
                if name == "[Content_Types].xml":
                    text = external_content_type_re.sub("", text)
                    text = calc_chain_content_type_re.sub("", text)
                content = text.encode("utf-8")
            zout.writestr(item, content)
    return target.getvalue()

def get_financial_year(month: int, year: int) -> str:
    if month >= 4:
        return f"{year}-{year + 1}"
    return f"{year - 1}-{year}"

def stable_scheme_key(scheme_name, asset_type=None, scheme_structure=None) -> str:
    return "|".join((
        norm_key(scheme_name),
        norm_key(asset_type),
        norm_key(scheme_structure),
    ))

def classify_macro_category(
    scheme_name,
    asset_type=None,
    scheme_structure=None,
    debt_equity=None,
    sales_prod_mis=None,
    fund_type=None,
) -> str:
    name = norm_key(scheme_name)
    asset = norm_key(asset_type)
    structure = norm_key(scheme_structure)
    debt_equity_key = norm_key(debt_equity)
    sales = norm_key(sales_prod_mis)
    detailed_type = norm_key(fund_type)

    category_aliases = {
        "active equity incl hybrid": "Active Equity (incl Hybrid)",
        "active equity": "Active Equity (incl Hybrid)",
        "hybrid": "Active Equity (incl Hybrid)",
        "arb esf": "Arb/ESF",
        "arbitrage esf": "Arb/ESF",
        "arbitrage": "Arb/ESF",
        "active debt incl hybrid": "Active Debt (incl Hybrid)",
        "active debt": "Active Debt (incl Hybrid)",
        "debt hybrid schemes": "Active Debt (incl Hybrid)",
        "cash": "Cash",
        "passive e d": "Passive-E/D",
        "passive equity": "Passive-E/D",
        "passive debt": "Passive-E/D",
        "index funds e d": "Passive-E/D",
        "interval schemes": "Interval Schemes",
    }
    if detailed_type in category_aliases:
        return category_aliases[detailed_type]
    if structure == "interval schemes" and asset in {"growth equity oriented schemes", "interval schemes"}:
        return "Interval Schemes"
    if name == "liquid fund":
        return "Cash"
    if sales == "arbitrage" or detailed_type == "arb esf":
        return "Arb/ESF"
    if (
        "etf" in name
        or "index" in name
        or "etf" in debt_equity_key
        or "index" in debt_equity_key
        or detailed_type in {"passive equity", "passive debt", "index funds e d"}
    ):
        return "Passive-E/D"
    if sales == "cash":
        return "Cash"
    if name in HYBRID_SCHEMES or "hybrid schemes" in asset:
        return "Active Debt (incl Hybrid)" if debt_equity_key == "debt" else "Active Equity (incl Hybrid)"
    if (
        detailed_type in {"active debt", "cash"}
        or "income debt" in asset
        or debt_equity_key == "debt"
        or sales in {"td", "fmp"}
    ):
        return "Cash" if detailed_type == "cash" or sales == "cash" else "Active Debt (incl Hybrid)"
    return "Active Equity (incl Hybrid)"

def _template_detail_bounds(ws) -> tuple[int, int]:
    grand_total_row = next(
        (
            row
            for row in range(DATA_START_ROW, ws.max_row + 1)
            if norm_key(ws.cell(row, 1).value) == "grand total"
        ),
        ws.max_row + 1,
    )
    detail_end = grand_total_row - 1
    while detail_end >= DATA_START_ROW and not ws.cell(detail_end, 1).value:
        detail_end -= 1
    return detail_end, grand_total_row

def _last_row_with_values(ws) -> int:
    max_col = ws.max_column
    for row in range(ws.max_row, 0, -1):
        if any(ws.cell(row, col).value not in (None, "") for col in range(1, max_col + 1)):
            return row
    return 1

@lru_cache(maxsize=1)
def template_catalog() -> tuple[dict, ...]:
    wb = load_workbook(TEMPLATE_PATH, data_only=True, keep_links=False)
    ws = wb[SHEET_FLAT] if SHEET_FLAT in wb.sheetnames else wb.worksheets[1]
    detail_end, _ = _template_detail_bounds(ws)
    catalog = []
    for row in range(DATA_START_ROW, detail_end + 1):
        scheme_name = clean_scheme(ws.cell(row, 1).value)
        if not scheme_name:
            continue
        fund_type = cell_value(ws, row, 6)
        is_grouped_member = ws.row_dimensions[row].outlineLevel > 0
        if (
            is_aggregate_scheme(scheme_name)
            or norm_key(scheme_name) in {norm_key(name) for name in CATEGORY_NAMES}
            or not fund_type
            or (ws.row_dimensions[row].outlineLevel == 0 and not is_grouped_member)
        ):
            continue
        entry = {
            "template_row": row,
            "scheme_name": scheme_name,
            "asset_type": cell_value(ws, row, 2),
            "scheme_structure": cell_value(ws, row, 3),
            "debt_equity": cell_value(ws, row, 4),
            "sales_prod_mis": cell_value(ws, row, 5),
            "fund_type": fund_type,
        }
        entry["scheme_key"] = stable_scheme_key(
            entry["scheme_name"], entry["asset_type"], entry["scheme_structure"]
        )
        entry["category"] = classify_macro_category(**{
            key: entry[key]
            for key in (
                "scheme_name",
                "asset_type",
                "scheme_structure",
                "debt_equity",
                "sales_prod_mis",
                "fund_type",
            )
        })
        catalog.append(entry)
    wb.close()
    return tuple(catalog)

def match_catalog_entry(
    scheme_name,
    asset_type=None,
    scheme_structure=None,
    debt_equity=None,
    sales_prod_mis=None,
) -> dict | None:
    if norm_key(scheme_name) == "elss" and "close ended" in norm_key(scheme_structure):
        for entry in template_catalog():
            if norm_key(entry["scheme_name"]) == "elss close ended":
                return entry

    candidates = [
        entry for entry in template_catalog()
        if norm_key(entry["scheme_name"]) == norm_key(scheme_name)
    ]
    if not candidates:
        return None
    if len(candidates) == 1:
        return candidates[0]

    inputs = (asset_type, scheme_structure, debt_equity, sales_prod_mis)
    fields = ("asset_type", "scheme_structure", "debt_equity", "sales_prod_mis")
    return max(
        candidates,
        key=lambda entry: sum(
            1
            for field, value in zip(fields, inputs)
            if value and norm_key(entry[field]) == norm_key(value)
        ),
    )

def enrich_scheme_record(record: dict) -> dict:
    matched = match_catalog_entry(
        record.get("scheme"),
        record.get("asset_type"),
        record.get("scheme_structure"),
        record.get("debt_equity"),
        record.get("sales_prod_mis"),
    )
    if matched:
        for field in (
            "asset_type",
            "scheme_structure",
            "debt_equity",
            "sales_prod_mis",
            "fund_type",
        ):
            if not record.get(field):
                record[field] = matched.get(field)
        record["scheme"] = matched.get("scheme_name") or record.get("scheme")
    record["scheme_key"] = stable_scheme_key(
        record.get("scheme"), record.get("asset_type"), record.get("scheme_structure")
    )
    record["fund_type"] = record.get("fund_type") or (
        matched.get("fund_type") if matched else None
    )
    return record

def process_upload_db(upload_bytes: bytes, filename: str) -> tuple[str, list[str]]:
    try:
        rows, month_info, warnings = parse_upload(upload_bytes, filename)
    except Exception as e:
        raise ValueError(f"Failed to parse uploaded Excel file: {str(e)}")

    if not rows:
        raise ValueError("Invalid sheet structure: No scheme rows or metric columns detected.")

    month = month_info["month"]
    year = month_info["year"]
    fy = get_financial_year(month, year)

    _UPLOAD_LOCK.acquire()
    conn = None
    try:
        conn = get_db_connection()
        conn.execute("BEGIN IMMEDIATE")
        cursor = conn.cursor()

        if month == 4:
            cursor.execute("SELECT COUNT(*) FROM amfi_metrics WHERE financial_year = ?", (fy,))
            fy_exists = cursor.fetchone()[0] > 0

            if not fy_exists:
                prev_fy = f"{year - 1}-{year}"
                cursor.execute("""
                    SELECT scheme_key, scheme_name, asset_type, scheme_structure, debt_equity,
                           sales_prod_mis, fund_type,
                           no_schemes, folios, funds_mobilized, redemption, net_inflow,
                           aum, avg_aum, seg_portfolios, seg_aum
                    FROM amfi_metrics
                    WHERE month = 3 AND year = ? AND financial_year = ?
                """, (year, prev_fy))
                march_records = cursor.fetchall()

                if not march_records:
                    cursor.execute("""
                        SELECT scheme_key, scheme_name, asset_type, scheme_structure, debt_equity,
                               sales_prod_mis, fund_type,
                               no_schemes, folios, funds_mobilized, redemption, net_inflow,
                               aum, avg_aum, seg_portfolios, seg_aum
                        FROM amfi_metrics
                        WHERE month = 3 AND year = ?
                    """, (year,))
                    march_records = cursor.fetchall()

                baseline_records = [
                    (
                        r["scheme_key"], r["scheme_name"], r["asset_type"], r["scheme_structure"],
                        r["debt_equity"], r["sales_prod_mis"], r["fund_type"],
                        r["no_schemes"], r["folios"], r["funds_mobilized"], r["redemption"], r["net_inflow"],
                        r["aum"], r["avg_aum"], r["seg_portfolios"], r["seg_aum"],
                        3, year, fy
                    )
                    for r in march_records
                ]

                if baseline_records:
                    cursor.executemany("""
                        INSERT OR IGNORE INTO amfi_metrics (
                            scheme_key, scheme_name, asset_type, scheme_structure, debt_equity,
                            sales_prod_mis, fund_type,
                            no_schemes, folios, funds_mobilized, redemption, net_inflow,
                            aum, avg_aum, seg_portfolios, seg_aum,
                            month, year, financial_year
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, baseline_records)

        insert_records = [
            (
                record["scheme_key"],
                record["scheme"],
                record.get("asset_type"),
                record.get("scheme_structure"),
                record.get("debt_equity"),
                record.get("sales_prod_mis"),
                record.get("fund_type"),
                record["metrics"].get("no_schemes"),
                record["metrics"].get("folios"),
                record["metrics"].get("funds_mobilized"),
                record["metrics"].get("redemption"),
                record["metrics"].get("net_inflow"),
                record["metrics"].get("net_aum"),
                record["metrics"].get("avg_aum"),
                record["metrics"].get("seg_portfolios"),
                record["metrics"].get("seg_aum"),
                month,
                year,
                fy
            )
            for record in rows
        ]

        cursor.execute(
            """
            DELETE FROM amfi_metrics
            WHERE month = ? AND year = ? AND financial_year = ?
            """,
            (month, year, fy),
        )

        cursor.executemany("""
            INSERT OR REPLACE INTO amfi_metrics (
                scheme_key, scheme_name, asset_type, scheme_structure, debt_equity,
                sales_prod_mis, fund_type,
                no_schemes, folios, funds_mobilized, redemption, net_inflow,
                aum, avg_aum, seg_portfolios, seg_aum,
                month, year, financial_year, last_modified
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
        """, insert_records)

        validation_errors = validate_uploaded_month_against_generated(fy, month, year, rows, conn=conn)
        if validation_errors:
            preview = "; ".join(validation_errors[:8])
            suffix = f" (+{len(validation_errors) - 8} more)" if len(validation_errors) > 8 else ""
            raise ValueError(f"Generated workbook reconciliation failed for {month_info['key']}: {preview}{suffix}")
        conn.commit()
        return month_info["key"], warnings
    except ValueError:
        if conn is not None:
            conn.rollback()
        raise
    except Exception as e:
        if conn is not None:
            conn.rollback()
        raise ValueError(f"Database error during ingestion: {str(e)}")
    finally:
        if conn is not None:
            conn.close()
        _UPLOAD_LOCK.release()

def extract_column_styles(ws, start_col: int, count: int) -> tuple[list[float], list[list[dict]]]:
    widths = []
    styles = []
    for col_offset in range(count):
        col = start_col + col_offset
        widths.append(ws.column_dimensions[get_column_letter(col)].width or 12.0)
        col_styles = []
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row, col)
            col_styles.append({
                "font": copy(cell.font) if cell.font else None,
                "fill": copy(cell.fill) if cell.fill else None,
                "border": copy(cell.border) if cell.border else None,
                "alignment": copy(cell.alignment) if cell.alignment else None,
                "number_format": cell.number_format,
                "value": cell.value if row <= 3 else None
            })
        styles.append(col_styles)
    return widths, styles

def apply_column_style(ws, col: int, width: float, styles: list[dict]):
    ws.column_dimensions[get_column_letter(col)].width = width
    for row_idx, style_info in enumerate(styles, 1):
        cell = ws.cell(row_idx, col)
        if style_info["font"]: cell.font = copy(style_info["font"])
        if style_info["fill"]: cell.fill = copy(style_info["fill"])
        if style_info["border"]: cell.border = copy(style_info["border"])
        if style_info["alignment"]: cell.alignment = copy(style_info["alignment"])
        cell.number_format = style_info["number_format"]
        if style_info["value"] is not None:
            cell.value = style_info["value"]

def write_form_column_subtotals(ws, col: int) -> None:
    col_letter = get_column_letter(col)
    for summary_row, row_ranges in _FORM_SUBTOTAL_ROWS.items():
        parts = [f"SUM({col_letter}{r1}:{col_letter}{r2})" for r1, r2 in row_ranges]
        ws.cell(summary_row, col).value = "=" + " + ".join(parts)

def get_month_seq(month: int, year: int, start_year: int) -> int:
    if month == 3 and year == start_year:
        return 0
    elif month >= 4 and year == start_year:
        return month - 3
    elif month <= 3 and year == start_year + 1:
        return month + 9
    return (year - start_year) * 12 + month

def clean_merged_ranges(ws, start_col: int):
    for r in list(ws.merged_cells.ranges):
        if r.min_col >= start_col or r.max_col >= start_col:
            ws.merged_cells.remove(r)

def copy_cell_style(source, target) -> None:
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.number_format = source.number_format
    target.protection = copy(source.protection)

def set_font_size(cell, size: float) -> None:
    font = copy(cell.font)
    font.sz = size
    cell.font = font

def set_font_name(cell, name: str) -> None:
    font = copy(cell.font)
    font.name = name
    cell.font = font

def _uses_generated_number_font(cell) -> bool:
    value = cell.value
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str) and value.startswith("="):
        return True
    return hasattr(value, "text")

def _normalize_generated_number_fonts(wb) -> None:
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if _uses_generated_number_font(cell):
                    set_font_name(cell, "Calibri")

def _record_key(record: dict) -> str:
    if record.get("scheme_key"):
        return record["scheme_key"]
    matched = match_catalog_entry(
        record.get("scheme_name"),
        record.get("asset_type"),
        record.get("scheme_structure"),
        record.get("debt_equity"),
        record.get("sales_prod_mis"),
    )
    return (
        (matched.get("scheme_key") if matched else None)
        or stable_scheme_key(
            record.get("scheme_name"),
            record.get("asset_type"),
            record.get("scheme_structure"),
        )
    )

def _flat_record_key(record: dict) -> str:
    if _is_fof_domestic_scheme(record.get("scheme_name")):
        return stable_scheme_key("Fund of Funds Scheme (Domestic)", "", "")
    if (
        norm_key(record.get("asset_type")) == "interval schemes"
        and norm_key(record.get("scheme_structure")) == "interval schemes"
        and norm_key(record.get("scheme_name")) in {"income debt oriented schemes", "growth equity oriented schemes", "other schemes"}
    ):
        return stable_scheme_key(record.get("scheme_name"), "Interval Schemes", "-")
    matched = match_catalog_entry(
        record.get("scheme_name"),
        record.get("asset_type"),
        record.get("scheme_structure"),
        record.get("debt_equity"),
        record.get("sales_prod_mis"),
    )
    return (matched.get("scheme_key") if matched else None) or _record_key(record)

def _merge_flat_record(existing: dict | None, record: dict, key: str) -> dict:
    if not existing:
        merged = dict(record)
        merged["scheme_key"] = key
        return merged
    for field in (
        "no_schemes",
        "folios",
        "funds_mobilized",
        "redemption",
        "net_inflow",
        "aum",
        "avg_aum",
        "seg_portfolios",
        "seg_aum",
    ):
        existing[field] = (existing.get(field) or 0) + (record.get(field) or 0)
    return existing

def _is_reference_only_scheme(name) -> bool:
    key = norm_key(name)
    return key.startswith("fund of funds scheme domestic") or key.startswith("data in respect")

def _is_fof_domestic_scheme(name) -> bool:
    return norm_key(name).startswith("fund of funds scheme domestic")

def _metric_from_record(record: dict | None, metric: str):
    if not record:
        return None
    db_key = "aum" if metric == "net_aum" else metric
    return record.get(db_key)

def _metric_display_value(record: dict | None, metric: str):
    value = _metric_from_record(record, metric)
    if (
        metric == "no_schemes"
        and value is not None
        and record
        and _is_reference_only_scheme(record.get("scheme_name"))
    ):
        numeric = number_or_none(value)
        if numeric is not None and float(numeric).is_integer():
            return f"##{int(numeric)}"
        return f"##{value}"
    return value

def _blank_interval_baseline_metric(record: dict | None) -> bool:
    if not record:
        return False
    if record.get("month") != 3 or norm_key(record.get("asset_type")) != "interval schemes":
        return False
    return all(
        number_or_none(record.get("aum" if metric == "net_aum" else metric)) in (None, 0)
        for metric in METRIC_ORDER
    )

def _sum_values(values) -> float:
    return round(sum(float(value) for value in values if isinstance(value, (int, float))), 6)

def _summary_values(scheme_key: str, maps: dict[int, dict[str, dict]], sorted_seqs: list[int]) -> tuple[float, float, float]:
    latest_seq = sorted_seqs[-1]
    latest = maps.get(latest_seq, {}).get(scheme_key)
    latest_aum = _metric_from_record(latest, "net_aum") or 0
    monthly_sales = (_metric_from_record(latest, "funds_mobilized") or 0) if latest_seq > 0 else 0
    fytd_sales = _sum_values(
        _metric_from_record(maps.get(seq, {}).get(scheme_key), "funds_mobilized")
        for seq in sorted_seqs
        if seq > 0
    )
    return latest_aum, monthly_sales, fytd_sales

def _add_styled_column(ws, style_ws, target_col: int, source_col: int) -> None:
    source_letter = get_column_letter(source_col)
    target_letter = get_column_letter(target_col)
    ws.column_dimensions[target_letter].width = style_ws.column_dimensions[source_letter].width
    for row in range(1, HEADER_ROW + 1):
        copy_cell_style(style_ws.cell(row, source_col), ws.cell(row, target_col))

def _flat_separator_source_col(style_ws, fallback_col: int) -> int:
    candidates = []
    for col in range(1, style_ws.max_column + 1):
        if style_ws.cell(2, col).value == "-" and style_ws.cell(3, col).value == "-":
            width = style_ws.column_dimensions[get_column_letter(col)].width
            if width is None or width <= 3:
                candidates.append(col)
    if not candidates:
        return fallback_col
    return min(candidates, key=lambda col: abs(col - fallback_col))

def _entry_from_record(record: dict) -> dict:
    entry = {
        "template_row": DATA_START_ROW,
        "scheme_key": _record_key(record),
        "scheme_name": record.get("scheme_name"),
        "asset_type": record.get("asset_type"),
        "scheme_structure": record.get("scheme_structure"),
        "debt_equity": record.get("debt_equity"),
        "sales_prod_mis": record.get("sales_prod_mis"),
        "fund_type": record.get("fund_type"),
    }
    entry["category"] = classify_macro_category(**{
        key: entry.get(key)
        for key in (
            "scheme_name",
            "asset_type",
            "scheme_structure",
            "debt_equity",
            "sales_prod_mis",
            "fund_type",
        )
    })
    return entry

def _build_output_specs(sorted_seqs: list[int], records_by_seq: dict[int, list[dict]], start_year: int) -> list[dict]:
    specs = []
    source_col = 7

    def add(kind: str, **kwargs):
        nonlocal source_col
        specs.append({"kind": kind, "source_col": source_col, **kwargs})
        source_col += 1

    if 0 in sorted_seqs:
        for metric in METRIC_ORDER:
            add("metric", seq=0, metric=metric)

    positive_seqs = [value for value in sorted_seqs if value > 0]
    latest_positive_seq = positive_seqs[-1] if positive_seqs else None
    for seq in positive_seqs:
        if seq != 11:
            add("separator")
        for offset, metric in enumerate(METRIC_ORDER):
            add("metric", seq=seq, metric=metric)
        if seq >= 10:
            add("separator")
            add("cumulative", seq=seq, metric="funds_mobilized")
            add("cumulative", seq=seq, metric="net_inflow")
            add("separator")
            add("growth", seq=seq, metric="funds_mobilized")
            add("growth", seq=seq, metric="net_inflow")
            add("growth", seq=seq, metric="net_aum")

    if positive_seqs and latest_positive_seq < 10:
        add("separator")
        add("cumulative", seq=latest_positive_seq, metric="funds_mobilized")
        add("cumulative", seq=latest_positive_seq, metric="net_inflow")
        add("separator")
        add("growth", seq=latest_positive_seq, metric="funds_mobilized")
        add("growth", seq=latest_positive_seq, metric="net_inflow")
        add("growth", seq=latest_positive_seq, metric="net_aum")
    return specs

def _spec_value(spec: dict, scheme_key: str, maps: dict[int, dict[str, dict]], sorted_seqs: list[int], summary: tuple) -> object:
    kind = spec["kind"]
    if kind == "separator":
        return "-"
    if kind == "metric":
        record = maps.get(spec["seq"], {}).get(scheme_key)
        if spec["seq"] == 0 and _blank_interval_baseline_metric(record):
            return None
        return _metric_from_record(record, spec["metric"])
    if kind == "cumulative":
        return _sum_values(
            _metric_from_record(maps.get(seq, {}).get(scheme_key), spec["metric"])
            for seq in sorted_seqs
            if 0 < seq <= spec["seq"]
        )
    if kind == "growth":
        current_seq = spec["seq"]
        previous_candidates = [seq for seq in sorted_seqs if seq < current_seq]
        previous_seq = previous_candidates[-1] if previous_candidates else 0
        previous = _metric_from_record(maps.get(previous_seq, {}).get(scheme_key), spec["metric"]) or 0
        current = _metric_from_record(maps.get(current_seq, {}).get(scheme_key), spec["metric"]) or 0
        return 0 if previous == 0 else (current - previous) / previous
    return None

def _spec_column_maps(specs: list[dict]) -> tuple[dict[tuple[int, str], int], dict[str, list[int]]]:
    metric_cols = {}
    positive_metric_cols = defaultdict(list)
    for col, spec in enumerate(specs, start=7):
        if spec["kind"] == "metric":
            key = (spec["seq"], spec["metric"])
            metric_cols[key] = col
            if spec["seq"] > 0:
                positive_metric_cols[spec["metric"]].append(col)
    return metric_cols, positive_metric_cols

def _growth_formula_for_row(spec: dict, row: int, sorted_seqs: list[int], metric_cols: dict[tuple[int, str], int]) -> str:
    current_seq = spec["seq"]
    previous_candidates = [seq for seq in sorted_seqs if seq < current_seq]
    previous_seq = previous_candidates[-1] if previous_candidates else 0
    previous_col = metric_cols.get((previous_seq, spec["metric"]))
    current_col = metric_cols.get((current_seq, spec["metric"]))
    if not previous_col or not current_col:
        return '="-"'
    previous_ref = f"{get_column_letter(previous_col)}{row}"
    current_ref = f"{get_column_letter(current_col)}{row}"
    return f'=IFERROR(({current_ref}-{previous_ref})/{previous_ref},"-")'

def _growth_formula_with_previous_value(
    spec: dict,
    row: int,
    sorted_seqs: list[int],
    metric_cols: dict[tuple[int, str], int],
    previous_value=None,
) -> str:
    formula = _growth_formula_for_row(spec, row, sorted_seqs, metric_cols)
    if formula != '="-"':
        return formula
    current_col = metric_cols.get((spec["seq"], spec["metric"]))
    previous_number = number_or_none(previous_value)
    if not current_col or previous_number in (None, 0):
        return '="-"'
    current_ref = f"{get_column_letter(current_col)}{row}"
    return f'=IFERROR(({current_ref}-{previous_number})/{previous_number},"-")'

def _apply_growth_number_format(cell) -> None:
    cell.number_format = "0%"

def _previous_metric_value_for_key(spec: dict, sorted_seqs: list[int], maps: dict[int, dict], scheme_key: str):
    previous_candidates = [seq for seq in sorted_seqs if seq < spec["seq"]]
    if not previous_candidates:
        return None
    previous_seq = previous_candidates[-1]
    return _metric_from_record(maps.get(previous_seq, {}).get(scheme_key), spec["metric"])

def _previous_metric_sum_for_keys(spec: dict, sorted_seqs: list[int], maps: dict[int, dict], scheme_keys: list[str]):
    values = []
    for scheme_key in scheme_keys:
        value = _previous_metric_value_for_key(spec, sorted_seqs, maps, scheme_key)
        if value is not None:
            values.append(value)
    return _sum_values(values) if values else None

def _cumulative_formula_for_row(spec: dict, row: int, positive_metric_cols: dict[str, list[int]], specs: list[dict]) -> str:
    refs = [
        f"{get_column_letter(col)}{row}"
        for col, source_spec in enumerate(specs, start=7)
        if (
            source_spec["kind"] == "metric"
            and source_spec["metric"] == spec["metric"]
            and 0 < source_spec["seq"] <= spec["seq"]
        )
    ]
    return f'=SUM({",".join(refs)})' if refs else "=0"

def _set_dynamic_headers(ws, specs: list[dict], records_by_seq: dict[int, list[dict]], sorted_seqs: list[int], start_year: int) -> None:
    latest_seq = sorted_seqs[-1]
    latest_record = records_by_seq[latest_seq][0]
    latest_key = f"{MONTH_ABBR[latest_record['month']]}'{str(latest_record['year'])[-2:]}"
    metric_columns_by_seq = defaultdict(list)
    col = 7
    for spec in specs:
        kind = spec["kind"]
        if kind == "separator":
            ws.cell(1, col).value = None
            ws.cell(2, col).value = "-"
            ws.cell(3, col).value = "-"
        elif kind == "metric":
            record = records_by_seq[spec["seq"]][0]
            key = f"{MONTH_ABBR[record['month']]}'{str(record['year'])[-2:]}"
            metric_columns_by_seq[spec["seq"]].append(col)
            label_info = month_info(record["month"], record["year"])
            ws.cell(3, col).value = metric_label(spec["metric"], label_info)
            if spec["metric"] == METRIC_ORDER[0]:
                ws.cell(2, col).value = key
                ws.cell(1, col).value = None
                ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 8)
                ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 8)
                ws.cell(1, col).alignment = copy(ws.cell(1, col).alignment)
                ws.cell(1, col).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(2, col).alignment = copy(ws.cell(2, col).alignment)
                ws.cell(2, col).alignment = Alignment(horizontal="center", vertical="center")
        elif kind == "cumulative":
            ws.cell(3, col).value = (
                "Funds Mobilized  (INR in crore)"
                if spec["metric"] == "funds_mobilized"
                else "Net Inflow (+ve)/Outflow (-ve)  (INR in crore)"
            )
            if spec["metric"] == "funds_mobilized":
                current = records_by_seq[spec["seq"]][0]
                current_key = f"{MONTH_ABBR[current['month']]}'{str(current['year'])[-2:]}"
                ws.cell(1, col).value = None
                ws.cell(2, col).value = f"Apr'{str(start_year)[-2:]} to {current_key}"
                ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
                ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 1)
        elif kind == "growth":
            previous_candidates = [seq for seq in sorted_seqs if seq < spec["seq"]]
            previous_seq = previous_candidates[-1] if previous_candidates else 0
            previous = records_by_seq[previous_seq][0]
            current = records_by_seq[spec["seq"]][0]
            previous_key = f"{MONTH_ABBR[previous['month']]}'{str(previous['year'])[-2:]}"
            current_key = f"{MONTH_ABBR[current['month']]}'{str(current['year'])[-2:]}"
            labels = {
                "funds_mobilized": f"GS - Growth-({previous_key} Vs {current_key})",
                "net_inflow": f"NS - Growth-({previous_key} Vs {current_key})",
                "net_aum": f"AUM - Growth-({previous_key} Vs {current_key})",
            }
            ws.cell(1, col).value = None
            ws.cell(2, col).value = "-"
            ws.cell(3, col).value = labels[spec["metric"]]
        col += 1

def _rebuild_flat_sheet(wb, style_wb, records_by_seq: dict[int, list[dict]], sorted_seqs: list[int], start_year: int):
    ws = wb[SHEET_FLAT] if SHEET_FLAT in wb.sheetnames else wb.worksheets[1]
    style_ws = style_wb[SHEET_FLAT] if SHEET_FLAT in style_wb.sheetnames else style_wb.worksheets[1]
    style_max_col = style_ws.max_column
    catalog = template_catalog()
    catalog_keys = {entry["scheme_key"] for entry in catalog}
    maps = {}
    for seq, records in records_by_seq.items():
        seq_map = {}
        for record in records:
            key = _flat_record_key(record)
            seq_map[key] = _merge_flat_record(seq_map.get(key), record, key)
        maps[seq] = seq_map

    extra_by_key = {}
    for seq in sorted_seqs:
        for record in records_by_seq[seq]:
            key = _flat_record_key(record)
            if key not in catalog_keys and not _is_reference_only_scheme(record.get("scheme_name")):
                extra_by_key[key] = _entry_from_record(record)
    entries = list(catalog) + list(extra_by_key.values())
    grouped = {category: [] for category in CATEGORY_NAMES}
    for entry in entries:
        grouped[entry["category"]].append(entry)

    specs = _build_output_specs(sorted_seqs, records_by_seq, start_year)
    metric_cols, positive_metric_cols = _spec_column_maps(specs)
    clean_merged_ranges(ws, 1)
    if ws.max_column > ATTRIBUTE_COLUMN_COUNT:
        ws.delete_cols(ATTRIBUTE_COLUMN_COUNT + 1, ws.max_column - ATTRIBUTE_COLUMN_COUNT)
    if ws.max_row >= DATA_START_ROW:
        ws.delete_rows(DATA_START_ROW, ws.max_row - DATA_START_ROW + 1)

    for index, spec in enumerate(specs, start=7):
        source_col = min(spec["source_col"], style_max_col)
        if spec["kind"] == "separator":
            source_col = _flat_separator_source_col(style_ws, source_col)
        _add_styled_column(ws, style_ws, index, source_col)
    _set_dynamic_headers(ws, specs, records_by_seq, sorted_seqs, start_year)

    detail_end, grand_total_source = _template_detail_bounds(style_ws)
    footer_end = _last_row_with_values(style_ws)
    source_rows_by_label = {
        norm_key(style_ws.cell(row, 1).value): row
        for row in range(DATA_START_ROW, footer_end + 1)
        if style_ws.cell(row, 1).value
    }
    row_cursor = DATA_START_ROW
    detail_rows = []
    subtotal_rows = []
    for category in CATEGORY_NAMES:
        category_row = row_cursor
        source_category_row = source_rows_by_label.get(norm_key(category), grand_total_source)
        ws.cell(category_row, 1).value = category
        for col in range(1, ws.max_column + 1):
            source_col = col if col <= 6 else specs[col - 7]["source_col"]
            copy_cell_style(
                style_ws.cell(source_category_row, min(source_col, style_max_col)),
                ws.cell(category_row, col),
            )
            ws.cell(category_row, col).value = None
        ws.cell(category_row, 1).value = category
        row_cursor += 1

        children = []
        for entry in grouped[category]:
            target_row = row_cursor
            source_row = min(entry.get("template_row") or DATA_START_ROW, detail_end)
            attrs = (
                entry["scheme_name"],
                entry.get("asset_type"),
                entry.get("scheme_structure"),
                entry.get("debt_equity"),
                entry.get("sales_prod_mis"),
                entry.get("fund_type"),
            )
            for col, value in enumerate(attrs, start=1):
                copy_cell_style(style_ws.cell(source_row, col), ws.cell(target_row, col))
                ws.cell(target_row, col).value = value

            summary = _summary_values(entry["scheme_key"], maps, sorted_seqs)
            values = {}
            for col, spec in enumerate(specs, start=7):
                copy_cell_style(
                    style_ws.cell(source_row, min(spec["source_col"], style_max_col)),
                    ws.cell(target_row, col),
                )
                if spec["kind"] == "growth":
                    previous_value = _previous_metric_value_for_key(spec, sorted_seqs, maps, entry["scheme_key"])
                    value = _growth_formula_with_previous_value(spec, target_row, sorted_seqs, metric_cols, previous_value)
                    _apply_growth_number_format(ws.cell(target_row, col))
                elif spec["kind"] == "cumulative":
                    value = _cumulative_formula_for_row(spec, target_row, positive_metric_cols, specs)
                else:
                    value = _spec_value(spec, entry["scheme_key"], maps, sorted_seqs, summary)
                ws.cell(target_row, col).value = value
            children.append(target_row)
            detail_rows.append(target_row)
            row_cursor += 1

        subtotal_row = row_cursor
        subtotal_rows.append(subtotal_row)
        source_subtotal_row = source_rows_by_label.get(norm_key(f"Total for {category}"), grand_total_source)
        for col in range(1, ws.max_column + 1):
            source_col = col if col <= 6 else specs[col - 7]["source_col"]
            copy_cell_style(
                style_ws.cell(source_subtotal_row, min(source_col, style_max_col)),
                ws.cell(subtotal_row, col),
            )
            ws.cell(subtotal_row, col).value = None
        ws.cell(subtotal_row, 1).value = f"Total for {category}"
        if children:
            ws.row_dimensions.group(children[0], children[-1], outline_level=1, hidden=True)
            ws.row_dimensions[category_row].collapsed = True
            first_child = children[0]
            last_child = children[-1]
            child_keys = [entry["scheme_key"] for entry in grouped[category]]
            for col, spec in enumerate(specs, start=7):
                cell = ws.cell(subtotal_row, col)
                if spec["kind"] == "separator":
                    cell.value = None
                elif spec["kind"] == "growth":
                    previous_value = _previous_metric_sum_for_keys(spec, sorted_seqs, maps, child_keys)
                    cell.value = _growth_formula_with_previous_value(spec, subtotal_row, sorted_seqs, metric_cols, previous_value)
                    _apply_growth_number_format(cell)
                else:
                    col_letter = get_column_letter(col)
                    cell.value = f"=SUM({col_letter}{first_child}:{col_letter}{last_child})"
        row_cursor += 1

    last_detail_row = row_cursor - 1
    footer_start = grand_total_source
    fof_records = [
        record
        for record in records_by_seq[sorted_seqs[-1]]
        if _is_reference_only_scheme(record.get("scheme_name"))
    ]
    fof_record = fof_records[0] if fof_records else None
    fof_key = _flat_record_key(fof_record) if fof_record else ""
    fof_records_by_seq = {
        seq: next(
            (record for record in records_by_seq[seq] if _is_reference_only_scheme(record.get("scheme_name"))),
            None,
        )
        for seq in sorted_seqs
    }
    detail_keys = [entry["scheme_key"] for entry in entries]

    for source_row in range(footer_start, footer_end + 1):
        target_row = row_cursor
        label = norm_key(style_ws.cell(source_row, 1).value)
        for col in range(1, ws.max_column + 1):
            source_col = col if col <= 6 else specs[col - 7]["source_col"]
            source_cell = style_ws.cell(source_row, min(source_col, style_max_col))
            target_cell = ws.cell(target_row, col)
            copy_cell_style(source_cell, target_cell)
            if not (isinstance(source_cell.value, str) and source_cell.value.startswith("=")):
                target_cell.value = source_cell.value

        if label == "grand total":
            for col, spec in enumerate(specs, start=7):
                cell = ws.cell(target_row, col)
                if spec["kind"] == "separator":
                    cell.value = None
                elif spec["kind"] == "growth":
                    previous_value = _previous_metric_sum_for_keys(spec, sorted_seqs, maps, detail_keys)
                    cell.value = _growth_formula_with_previous_value(spec, target_row, sorted_seqs, metric_cols, previous_value)
                    _apply_growth_number_format(cell)
                else:
                    col_letter = get_column_letter(col)
                    refs = ",".join(f"{col_letter}{row}" for row in subtotal_rows)
                    cell.value = f"=SUM({refs})"
        elif label.startswith("fund of funds scheme domestic") and fof_record:
            summary = _summary_values(fof_key, maps, sorted_seqs)
            for col, spec in enumerate(specs, start=7):
                set_font_size(ws.cell(target_row, col), 10)
                if spec["kind"] == "growth":
                    previous_value = _previous_metric_value_for_key(spec, sorted_seqs, maps, fof_key)
                    value = _growth_formula_with_previous_value(spec, target_row, sorted_seqs, metric_cols, previous_value)
                    _apply_growth_number_format(ws.cell(target_row, col))
                elif spec["kind"] == "cumulative":
                    value = _cumulative_formula_for_row(spec, target_row, positive_metric_cols, specs)
                elif spec["kind"] == "metric":
                    value = _metric_display_value(fof_records_by_seq.get(spec["seq"]), spec["metric"])
                else:
                    value = _spec_value(spec, fof_key, maps, sorted_seqs, summary)
                ws.cell(target_row, col).value = value
        elif style_ws.cell(source_row, 1).value == "-":
            pass
        elif label == "total":
            grand_total_row = next(
                row for row in range(DATA_START_ROW, target_row)
                if norm_key(ws.cell(row, 1).value) == "grand total"
            )
            for col, spec in enumerate(specs, start=7):
                ws.cell(target_row, col).value = None if spec["kind"] == "separator" else f"={get_column_letter(col)}{grand_total_row}"
        elif label.startswith("data in respect") or label.startswith("include nfo"):
            ws.merge_cells(start_row=target_row, start_column=1, end_row=target_row, end_column=ws.max_column)
        row_cursor += 1

    if ws.max_row >= row_cursor:
        ws.delete_rows(row_cursor, ws.max_row - row_cursor + 1)
    ws.sheet_properties.outlinePr = Outline(summaryBelow=False, showOutlineSymbols=True)
    ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(ws.max_column)}{last_detail_row}"
    ws.freeze_panes = "B4"
    ws.sheet_view.showGridLines = style_ws.sheet_view.showGridLines
    return ws

def _rebuild_form_sheet(wb, style_wb, records_by_seq: dict[int, list[dict]], sorted_seqs: list[int]):
    ws = wb[SHEET_FORM] if SHEET_FORM in wb.sheetnames else wb.worksheets[2]
    style_ws = style_wb[SHEET_FORM] if SHEET_FORM in style_wb.sheetnames else style_wb.worksheets[2]
    march_widths, march_styles = extract_column_styles(style_ws, 3, 2)
    sep_widths, sep_styles = extract_column_styles(style_ws, 5, 1)
    metric_widths, metric_styles = extract_column_styles(style_ws, 6, 11)
    comp_sep_widths, comp_sep_styles = extract_column_styles(style_ws, 125, 1)
    cum_widths, cum_styles = extract_column_styles(style_ws, 126, 2)
    growth_sep_widths, growth_sep_styles = extract_column_styles(style_ws, 128, 1)
    growth_widths, growth_styles = extract_column_styles(style_ws, 129, 3)
    clean_merged_ranges(ws, 3)
    if ws.max_column > 2:
        ws.delete_cols(3, ws.max_column - 2)

    baseline_cols: dict[str, int] = {}
    month_cols: dict[int, dict[str, int]] = {}
    form_row_keys = _form_row_keys(ws)
    for seq in sorted_seqs:
        records = records_by_seq[seq]
        record = records[0]
        key = f"{MONTH_ABBR[record['month']]}'{str(record['year'])[-2:]}"
        record_map = {_record_key(item): item for item in records}
        fof_record = next(
            (item for item in records if _is_reference_only_scheme(item.get("scheme_name"))),
            None,
        )
        if seq == 0:
            start = ws.max_column + 1
            for offset in range(2):
                apply_column_style(ws, start + offset, march_widths[offset], march_styles[offset])
            info = month_info(record["month"], record["year"])
            ws.cell(MONTH_ROW, start).value = key
            ws.cell(HEADER_ROW, start).value = metric_label("net_aum", info)
            ws.cell(HEADER_ROW, start + 1).value = metric_label("avg_aum", info)
            baseline_cols = {"net_aum": start, "avg_aum": start + 1}
            ws.merge_cells(start_row=2, start_column=start, end_row=2, end_column=start + 1)
            for row in range(DATA_START_ROW, ws.max_row + 1):
                if row in _FORM_SUBTOTAL_ROWS:
                    write_form_column_subtotals(ws, start)
                    write_form_column_subtotals(ws, start + 1)
                    continue
                row_key = form_row_keys.get(row)
                matched = fof_record if row_key == "__fof_domestic__" else record_map.get(row_key)
                if matched:
                    ws.cell(row, start).value = matched.get("aum")
                    ws.cell(row, start + 1).value = matched.get("avg_aum")
            continue

        sep = ws.max_column + 1
        apply_column_style(ws, sep, sep_widths[0], sep_styles[0])
        ws.cell(MONTH_ROW, sep).value = "-"
        ws.cell(HEADER_ROW, sep).value = "-"
        start = ws.max_column + 1
        for offset in range(11):
            apply_column_style(ws, start + offset, metric_widths[offset], metric_styles[offset])
        info = month_info(record["month"], record["year"])
        ws.cell(MONTH_ROW, start).value = f"Monthly Report for the month of {info['full_month']} {record['year']} "
        ws.cell(HEADER_ROW, start).value = "Sr "
        ws.cell(HEADER_ROW, start + 1).value = "Scheme Name "
        month_cols[seq] = {}
        for offset, metric in enumerate(METRIC_ORDER):
            metric_col = start + 2 + offset
            ws.cell(HEADER_ROW, metric_col).value = metric_label(metric, info)
            month_cols[seq][metric] = metric_col
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=start + 10)
        ws.merge_cells(start_row=2, start_column=start, end_row=2, end_column=start + 10)
        for row in range(DATA_START_ROW, ws.max_row + 1):
            if row in _FORM_SUBTOTAL_ROWS:
                for offset in range(9):
                    write_form_column_subtotals(ws, start + 2 + offset)
                continue
            row_key = form_row_keys.get(row)
            matched = fof_record if row_key == "__fof_domestic__" else record_map.get(row_key)
            if matched:
                ws.cell(row, start).value = ws.cell(row, 1).value
                ws.cell(row, start + 1).value = ws.cell(row, 2).value
                for offset, metric in enumerate(METRIC_ORDER):
                    ws.cell(row, start + 2 + offset).value = (
                        _metric_display_value(matched, metric)
                        if row_key == "__fof_domestic__"
                        else _metric_from_record(matched, metric)
                    )

    positive_seqs = [seq for seq in sorted_seqs if seq > 0]
    if positive_seqs:
        latest_seq = positive_seqs[-1]
        latest_record = records_by_seq[latest_seq][0]
        latest_key = f"{MONTH_ABBR[latest_record['month']]}'{str(latest_record['year'])[-2:]}"

        sep = ws.max_column + 1
        apply_column_style(ws, sep, comp_sep_widths[0], comp_sep_styles[0])
        ws.cell(MONTH_ROW, sep).value = "-"
        ws.cell(HEADER_ROW, sep).value = "-"

        cum_start = ws.max_column + 1
        for offset in range(2):
            apply_column_style(ws, cum_start + offset, cum_widths[offset], cum_styles[offset])
        ws.cell(MONTH_ROW, cum_start).value = f"Apr'{str(latest_record['year'] if latest_record['month'] >= 4 else latest_record['year'] - 1)[-2:]} to {latest_key}"
        ws.cell(HEADER_ROW, cum_start).value = f"Funds Mobilized for the month of {month_info(latest_record['month'], latest_record['year'])['full_month']} {latest_record['year']} (INR in crore)"
        ws.cell(HEADER_ROW, cum_start + 1).value = f"Net Inflow (+ve)/Outflow (-ve) for the month of {month_info(latest_record['month'], latest_record['year'])['full_month']} {latest_record['year']} (INR in crore)"
        ws.merge_cells(start_row=2, start_column=cum_start, end_row=2, end_column=cum_start + 1)

        sep = ws.max_column + 1
        apply_column_style(ws, sep, growth_sep_widths[0], growth_sep_styles[0])
        ws.cell(MONTH_ROW, sep).value = "-"
        ws.cell(HEADER_ROW, sep).value = "-"

        growth_start = ws.max_column + 1
        for offset in range(3):
            apply_column_style(ws, growth_start + offset, growth_widths[offset], growth_styles[offset])
        previous_candidates = [seq for seq in sorted_seqs if seq < latest_seq]
        previous_seq = previous_candidates[-1] if previous_candidates else 0
        previous_record = records_by_seq[previous_seq][0]
        previous_key = f"{MONTH_ABBR[previous_record['month']]}'{str(previous_record['year'])[-2:]}"
        ws.cell(MONTH_ROW, growth_start).value = "-"
        ws.cell(MONTH_ROW, growth_start + 1).value = "-"
        ws.cell(MONTH_ROW, growth_start + 2).value = "-"
        ws.cell(HEADER_ROW, growth_start).value = f"GS - Growth-({previous_key} Vs {latest_key})"
        ws.cell(HEADER_ROW, growth_start + 1).value = f"NS - Growth-({previous_key} Vs {latest_key})"
        ws.cell(HEADER_ROW, growth_start + 2).value = f"AUM - Growth-({previous_key} Vs {latest_key})"

        funds_cols = [month_cols[seq]["funds_mobilized"] for seq in positive_seqs if seq in month_cols]
        net_cols = [month_cols[seq]["net_inflow"] for seq in positive_seqs if seq in month_cols]
        previous_cols = month_cols.get(previous_seq, {})
        if previous_seq == 0:
            previous_cols = {**previous_cols, **baseline_cols}
        latest_cols = month_cols.get(latest_seq, {})
        for row in range(DATA_START_ROW, ws.max_row + 1):
            ws.cell(row, cum_start).value = (
                f"=SUM({','.join(f'{get_column_letter(col)}{row}' for col in funds_cols)})"
                if funds_cols else "=0"
            )
            ws.cell(row, cum_start + 1).value = (
                f"=SUM({','.join(f'{get_column_letter(col)}{row}' for col in net_cols)})"
                if net_cols else "=0"
            )
            for offset, metric in enumerate(("funds_mobilized", "net_inflow", "net_aum")):
                previous_col = previous_cols.get(metric)
                current_col = latest_cols.get(metric)
                if previous_col and current_col:
                    previous_ref = f"{get_column_letter(previous_col)}{row}"
                    current_ref = f"{get_column_letter(current_col)}{row}"
                    ws.cell(row, growth_start + offset).value = f'=IFERROR(({current_ref}-{previous_ref})/{previous_ref},"-")'
                else:
                    ws.cell(row, growth_start + offset).value = '="-"'
                _apply_growth_number_format(ws.cell(row, growth_start + offset))
    return ws

def _form_row_keys(ws) -> dict[int, str]:
    keys = {}
    current_structure = None
    current_asset_type = None
    structure_serials = {"A", "B", "C"}
    section_serials = {"I", "II", "III", "IV", "V"}
    for row in range(DATA_START_ROW, ws.max_row + 1):
        raw_serial = str(ws.cell(row, 1).value or "").strip()
        name = clean_scheme(ws.cell(row, 2).value)
        if not name:
            continue
        name_key = norm_key(name)
        if raw_serial in structure_serials and (
            "open ended" in name_key
            or "close ended" in name_key
            or "interval" in name_key
        ):
            current_structure = name
            current_asset_type = None
            continue
        if is_aggregate_scheme(name):
            continue
        if _is_fof_domestic_scheme(name):
            keys[row] = "__fof_domestic__"
            continue
        if raw_serial in section_serials:
            if norm_key(current_structure) == "interval schemes":
                asset_type = current_structure
                structure = current_structure
                keys[row] = stable_scheme_key(name, asset_type, structure)
                continue
            if norm_key(current_structure) == "close ended schemes" and name_key == "other schemes":
                asset_type = name
                structure = current_structure
            else:
                current_asset_type = name
                continue
        else:
            asset_type = current_asset_type
            structure = current_structure
        matched = match_catalog_entry(name, asset_type, structure)
        keys[row] = (matched or {}).get("scheme_key") or stable_scheme_key(name, asset_type, structure)
    return keys

def _month_key_from_record(record: dict | None) -> str:
    if not record:
        return "-"
    return f"{MONTH_ABBR[record['month']]}'{str(record['year'])[-2:]}"

def _ns_growth(previous, current):
    previous_value = number_or_none(previous)
    current_value = number_or_none(current)
    if previous_value in (None, 0) or current_value is None:
        return None
    return (current_value - previous_value) / previous_value

def _ns_row(
    label: str,
    previous,
    current,
    row_number: int | None = None,
    previous_aum=None,
    current_aum=None,
) -> dict:
    previous_value = number_or_none(previous)
    current_value = number_or_none(current)
    previous_aum_value = number_or_none(previous_aum)
    current_aum_value = number_or_none(current_aum)
    return {
        "label": label,
        "row": row_number,
        "previous": 0.0 if previous_value is None else round(previous_value, 6),
        "current": 0.0 if current_value is None else round(current_value, 6),
        "growth": None if _ns_growth(previous_value, current_value) is None else round(_ns_growth(previous_value, current_value), 6),
        "previousAum": 0.0 if previous_aum_value is None else round(previous_aum_value, 6),
        "currentAum": 0.0 if current_aum_value is None else round(current_aum_value, 6),
        "aumGrowth": None if _ns_growth(previous_aum_value, current_aum_value) is None else round(_ns_growth(previous_aum_value, current_aum_value), 6),
    }

def _records_by_key(records_by_seq: dict[int, list[dict]], seq: int | None) -> dict[str, dict]:
    if seq is None:
        return {}
    return {_record_key(record): record for record in records_by_seq.get(seq, [])}

def _ns_total(records_by_seq: dict[int, list[dict]], seq: int | None, metric: str = "net_inflow") -> float:
    if seq is None:
        return 0.0
    db_metric = "aum" if metric == "net_aum" else metric
    return _sum_values(
        record.get(db_metric)
        for record in records_by_seq.get(seq, [])
        if not _is_reference_only_scheme(record.get("scheme_name"))
    )

def _ns_value(row_key: str, keyed_records: dict[str, dict], metric: str = "net_inflow") -> float:
    db_metric = "aum" if metric == "net_aum" else metric
    return float((keyed_records.get(row_key) or {}).get(db_metric) or 0)

def _ns_entry_rows(template_rows: tuple[tuple[int, int], ...]) -> list[tuple[int, dict]]:
    entries_by_row = {entry["template_row"]: entry for entry in template_catalog()}
    rows = []
    for ns_row, template_row in template_rows:
        entry = entries_by_row.get(template_row)
        if entry:
            rows.append((ns_row, entry))
    return rows

def build_ns_analysis_from_records(records_by_seq: dict[int, list[dict]], sorted_seqs: list[int]) -> dict:
    if not sorted_seqs:
        return {
            "previousMonth": "-",
            "currentMonth": "-",
            "total": _ns_row("TOTAL", 0, 0, 3),
            "headlineRows": [],
            "equityRows": [],
            "hybridRows": [],
        }

    current_seq = sorted_seqs[-1]
    previous_candidates = [seq for seq in sorted_seqs if seq < current_seq]
    previous_seq = previous_candidates[-1] if previous_candidates else None
    previous_records = _records_by_key(records_by_seq, previous_seq)
    current_records = _records_by_key(records_by_seq, current_seq)
    previous_month = _month_key_from_record(records_by_seq.get(previous_seq, [None])[0] if previous_seq is not None else None)
    current_month = _month_key_from_record(records_by_seq[current_seq][0])

    equity_rows = []
    for row_number, entry in _ns_entry_rows(NS_EQUITY_TEMPLATE_ROWS):
        key = entry["scheme_key"]
        equity_rows.append(_ns_row(
            entry["scheme_name"],
            _ns_value(key, previous_records),
            _ns_value(key, current_records),
            row_number,
            _ns_value(key, previous_records, "net_aum"),
            _ns_value(key, current_records, "net_aum"),
        ))

    hybrid_rows = []
    for row_number, entry in _ns_entry_rows(NS_HYBRID_TEMPLATE_ROWS):
        key = entry["scheme_key"]
        hybrid_rows.append(_ns_row(
            entry["scheme_name"],
            _ns_value(key, previous_records),
            _ns_value(key, current_records),
            row_number,
            _ns_value(key, previous_records, "net_aum"),
            _ns_value(key, current_records, "net_aum"),
        ))

    equity_total = _ns_row(
        "Equity excl Hybrid",
        _sum_values(row["previous"] for row in equity_rows),
        _sum_values(row["current"] for row in equity_rows),
        5,
        _sum_values(row["previousAum"] for row in equity_rows),
        _sum_values(row["currentAum"] for row in equity_rows),
    )
    hybrid_total = _ns_row(
        "Hybrid",
        _sum_values(row["previous"] for row in hybrid_rows),
        _sum_values(row["current"] for row in hybrid_rows),
        6,
        _sum_values(row["previousAum"] for row in hybrid_rows),
        _sum_values(row["currentAum"] for row in hybrid_rows),
    )
    equity_hybrid_total = _ns_row(
        "Equity + Hybrid",
        equity_total["previous"] + hybrid_total["previous"],
        equity_total["current"] + hybrid_total["current"],
        7,
        equity_total["previousAum"] + hybrid_total["previousAum"],
        equity_total["currentAum"] + hybrid_total["currentAum"],
    )
    hybrid_section_total = _ns_row(
        "Hybrid",
        hybrid_total["previous"],
        hybrid_total["current"],
        33,
        hybrid_total["previousAum"],
        hybrid_total["currentAum"],
    )
    equity_section_total = _ns_row(
        "EQUITY",
        equity_total["previous"],
        equity_total["current"],
        27,
        equity_total["previousAum"],
        equity_total["currentAum"],
    )

    return {
        "previousMonth": previous_month,
        "currentMonth": current_month,
        "total": _ns_row(
            "TOTAL",
            _ns_total(records_by_seq, previous_seq),
            _ns_total(records_by_seq, current_seq),
            3,
            _ns_total(records_by_seq, previous_seq, "net_aum"),
            _ns_total(records_by_seq, current_seq, "net_aum"),
        ),
        "headlineRows": [equity_total, hybrid_total, equity_hybrid_total],
        "equityRows": equity_rows,
        "equityTotal": equity_section_total,
        "hybridRows": hybrid_rows,
        "hybridTotal": hybrid_section_total,
    }

def _write_ns_analysis_row(ws, row: int, item: dict, formula_growth: bool = True) -> None:
    ws.cell(row, 1).value = item["label"]
    ws.cell(row, 2).value = item["previous"]
    ws.cell(row, 3).value = item["current"]
    ws.cell(row, 4).value = f'=IFERROR((C{row}-B{row})/B{row},"-")' if formula_growth else item.get("growth")
    ws.cell(row, 6).value = item.get("previousAum", 0)
    ws.cell(row, 7).value = item.get("currentAum", 0)
    ws.cell(row, 8).value = f'=IFERROR((G{row}-F{row})/F{row},"-")' if formula_growth else item.get("aumGrowth")

def _rebuild_ns_analysis_sheet(wb, style_wb, records_by_seq: dict[int, list[dict]], sorted_seqs: list[int]):
    if SHEET_NS in wb.sheetnames:
        ws = wb[SHEET_NS]
    else:
        ws = wb.create_sheet(SHEET_NS, 3)
    if SHEET_NS in style_wb.sheetnames:
        style_ws = style_wb[SHEET_NS]
        for col in range(1, min(style_ws.max_column, 8) + 1):
            ws.column_dimensions[get_column_letter(col)].width = style_ws.column_dimensions[get_column_letter(col)].width
            for row in range(1, style_ws.max_row + 1):
                copy_cell_style(style_ws.cell(row, col), ws.cell(row, col))
                if col == 1:
                    ws.row_dimensions[row].height = style_ws.row_dimensions[row].height
    if ws.max_column > 8:
        ws.delete_cols(9, ws.max_column - 8)

    analysis = build_ns_analysis_from_records(records_by_seq, sorted_seqs)
    ws.cell(1, 1).value = "AMFI"
    ws.cell(2, 1).value = "Net Sales"
    ws.cell(2, 2).value = analysis["previousMonth"]
    ws.cell(2, 3).value = analysis["currentMonth"]
    ws.cell(2, 4).value = "Growth %"
    ws.cell(1, 6).value = "AUM"
    ws.cell(2, 6).value = analysis["previousMonth"]
    ws.cell(2, 7).value = analysis["currentMonth"]
    ws.cell(2, 8).value = "Growth %"

    for row in range(3, max(ws.max_row, 33) + 1):
        for col in (2, 3, 4, 6, 7, 8):
            ws.cell(row, col).value = None

    _write_ns_analysis_row(ws, 3, analysis["total"])
    for item in analysis["headlineRows"]:
        _write_ns_analysis_row(ws, item["row"], item)
    ws.cell(9, 1).value = "Equity excl Hybrid"
    for col in (2, 3, 4, 6, 7, 8):
        ws.cell(9, col).value = "-"
    for item in analysis["equityRows"]:
        _write_ns_analysis_row(ws, item["row"], item)
    _write_ns_analysis_row(ws, 27, analysis["equityTotal"])

    ws.cell(29, 1).value = "Hybrid"
    ws.cell(29, 2).value = analysis["previousMonth"]
    ws.cell(29, 3).value = analysis["currentMonth"]
    ws.cell(29, 4).value = "Growth %"
    ws.cell(29, 6).value = analysis["previousMonth"]
    ws.cell(29, 7).value = analysis["currentMonth"]
    ws.cell(29, 8).value = "Growth %"
    for item in analysis["hybridRows"]:
        _write_ns_analysis_row(ws, item["row"], item)
    _write_ns_analysis_row(ws, 33, analysis["hybridTotal"])
    return ws

def _copy_row_style(ws, source_row: int, target_row: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, ws.max_column + 1):
        copy_cell_style(ws.cell(source_row, col), ws.cell(target_row, col))

def _sip_history_rows() -> list[dict]:
    if not SIP_HISTORY_PATH.exists():
        return []
    rows = []
    with SIP_HISTORY_PATH.open(newline="", encoding="utf-8") as handle:
        for row in csv.DictReader(handle):
            date_text = (row.get("date") or "").strip()
            if not date_text:
                continue
            try:
                date_value = datetime.strptime(date_text, "%Y-%m-%d")
            except ValueError:
                continue
            rows.append({
                "date": date_value,
                "outstanding_accounts": number_or_none(row.get("outstanding_accounts")),
                "new_registrations": number_or_none(row.get("new_registrations")),
                "discontinued": number_or_none(row.get("discontinued")),
                "contributing_accounts": number_or_none(row.get("contributing_accounts")),
                "contribution": number_or_none(row.get("contribution")),
                "aum": number_or_none(row.get("aum")),
            })
    return rows

def _sip_financial_year_label(date_value: datetime) -> str:
    start_year = date_value.year if date_value.month >= 4 else date_value.year - 1
    return f"FY {start_year}-{str(start_year + 1)[-2:]}"

def _sip_month_label(date_value: datetime) -> str:
    return "Jun" if date_value.month == 6 else calendar.month_name[date_value.month]

def _sip_annual_column(ws, fy_label: str) -> int | None:
    for col in range(2, ws.max_column + 1):
        if str(ws.cell(4, col).value or "").strip() == fy_label:
            return col
    return None

def _sip_month_row(ws, date_value: datetime) -> int | None:
    target = _sip_month_label(date_value).lower()
    for row in range(6, 18):
        if str(ws.cell(row, 1).value or "").strip().lower() == target:
            return row
    return None

def _sip_detail_row(ws, date_value: datetime) -> int:
    for row in range(22, ws.max_row + 1):
        value = ws.cell(row, 1).value
        if isinstance(value, datetime) and value.year == date_value.year and value.month == date_value.month:
            return row

    for row in range(22, ws.max_row + 1):
        if ws.cell(row, 1).value in (None, ""):
            return row
        if not isinstance(ws.cell(row, 1).value, datetime):
            ws.insert_rows(row)
            _copy_row_style(ws, row + 1, row)
            return row

    ws.append([None] * ws.max_column)
    _copy_row_style(ws, ws.max_row - 1, ws.max_row)
    return ws.max_row

def _sync_sip_sheet(wb) -> None:
    if SHEET_SIP not in wb.sheetnames:
        return
    ws = wb[SHEET_SIP]
    for row in _sip_history_rows():
        date_value = row["date"]
        contribution = row.get("contribution")

        annual_col = _sip_annual_column(ws, _sip_financial_year_label(date_value))
        annual_row = _sip_month_row(ws, date_value)
        if annual_col and annual_row and contribution is not None:
            ws.cell(annual_row, annual_col).value = contribution

        detail_row = _sip_detail_row(ws, date_value)
        values = [
            date_value,
            row.get("outstanding_accounts"),
            row.get("new_registrations"),
            row.get("discontinued"),
            row.get("contributing_accounts"),
            contribution,
            row.get("aum"),
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(detail_row, col).value = value

def _refresh_analytics_trend_helper(wb, records_by_seq: dict[int, list[dict]], sorted_seqs: list[int]) -> None:
    if "Analytics" not in wb.sheetnames:
        return
    ws = wb["Analytics"]
    ws.cell(26, 16).value = "AUM Trend Chart Source"
    ws.cell(27, 16).value = "Month"
    ws.cell(27, 17).value = "AUM"
    for col in (16, 17):
        copy_cell_style(ws.cell(27, col - 15), ws.cell(27, col))
    for row in range(28, 41):
        for col in (16, 17):
            cell = ws.cell(row, col)
            cell.value = None
            copy_cell_style(ws.cell(row, col - 15), cell)

    for index, seq in enumerate(sorted_seqs[:13], start=28):
        record = records_by_seq[seq][0]
        label = f"{MONTH_ABBR[record['month']]}'{str(record['year'])[-2:]}"
        total_aum = _sum_values(
            item.get("aum")
            for item in records_by_seq[seq]
            if not _is_reference_only_scheme(item.get("scheme_name"))
        )
        ws.cell(index, 16).value = label
        ws.cell(index, 17).value = total_aum
        ws.cell(index, 17).number_format = '#,##0'

    ws.cell(6, 12).value = "=COUNT($Q$28:$Q$40)"

def _retarget_sheet_references(wb, old_title: str, new_title: str) -> None:
    if old_title == new_title:
        return
    old_escaped = old_title.replace("'", "''")
    new_escaped = new_title.replace("'", "''")
    old_quoted = f"'{old_escaped}'!"
    new_quoted = f"'{new_escaped}'!"
    old_plain = f"{old_title}!"
    new_plain = f"{new_title}!"

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str):
                    if value.startswith("="):
                        cell.value = value.replace(old_quoted, new_quoted).replace(old_plain, new_plain)
                    else:
                        cell.value = value.replace(old_title, new_title)
                elif hasattr(value, "text") and isinstance(value.text, str):
                    value.text = value.text.replace(old_quoted, new_quoted).replace(old_plain, new_plain)
    for defined_name in wb.defined_names.values():
        text = getattr(defined_name, "attr_text", None)
        if isinstance(text, str):
            defined_name.attr_text = text.replace(old_quoted, new_quoted).replace(old_plain, new_plain).replace(old_title, new_title)

def _reset_sheet_view(ws) -> None:
    ws.sheet_view.topLeftCell = "A1"
    pane = ws.sheet_view.pane
    if not pane:
        ws.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]
        return

    pane.activePane = "bottomRight"
    top_left = pane.topLeftCell or "A1"
    match = re.match(r"([A-Z]+)(\d+)", top_left)
    if not match:
        ws.sheet_view.selection = [Selection(pane="bottomRight", activeCell="A1", sqref="A1")]
        return

    top_left_col, top_left_row = match.groups()
    selections = []
    if pane.xSplit:
        selections.append(Selection(pane="topRight", activeCell=f"{top_left_col}1", sqref=f"{top_left_col}1"))
    if pane.ySplit:
        selections.append(Selection(pane="bottomLeft", activeCell=f"A{top_left_row}", sqref=f"A{top_left_row}"))
    selections.append(Selection(pane="bottomRight", activeCell=top_left, sqref=top_left))
    ws.sheet_view.selection = selections

def compile_excel_for_fy(fy: str, conn: sqlite3.Connection | None = None) -> bytes:
    owns_connection = conn is None
    if conn is None:
        conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM amfi_metrics WHERE financial_year = ?", (fy,))
        records = [dict(row) for row in cursor.fetchall()]
        start_year = int(fy.split("-")[0])
    finally:
        if owns_connection:
            conn.close()

    if not records:
        raise ValueError(f"No records found in database for fiscal year: {fy}")

    records_by_seq = defaultdict(list)
    for record in records:
        records_by_seq[get_month_seq(record["month"], record["year"], start_year)].append(record)
    sorted_seqs = sorted(records_by_seq)

    wb = load_workbook(TEMPLATE_PATH, keep_links=False)
    style_wb = load_workbook(TEMPLATE_PATH, keep_links=False)
    ws_flat = _rebuild_flat_sheet(wb, style_wb, records_by_seq, sorted_seqs, start_year)
    ws_form = _rebuild_form_sheet(wb, style_wb, records_by_seq, sorted_seqs)
    _rebuild_ns_analysis_sheet(wb, style_wb, records_by_seq, sorted_seqs)
    _sync_sip_sheet(wb)
    _refresh_analytics_trend_helper(wb, records_by_seq, sorted_seqs)
    old_flat_title = ws_flat.title
    old_form_title = ws_form.title
    latest_record = records_by_seq[sorted_seqs[-1]][0]
    latest_key = f"{MONTH_ABBR[latest_record['month']]}'{str(latest_record['year'])[-2:]}"
    baseline_key = f"Mar'{str(start_year)[-2:]}"
    ws_flat.title = f"AMFI-{baseline_key} to {latest_key}"
    ws_form.title = f"AMFI-{baseline_key} to {latest_key}-AMFI form"
    _retarget_sheet_references(wb, old_flat_title, ws_flat.title)
    _retarget_sheet_references(wb, old_form_title, ws_form.title)
    wb.active = 0
    active_ws = wb.worksheets[0]
    for ws in wb.worksheets:
        # Only the active sheet may be tab-selected. The template ships with the
        # data sheet flagged tabSelected, which — combined with wb.active=0 —
        # makes Excel open the sheets as a group (both tabs highlighted) and
        # disables the outline collapse/expand buttons until the user ungroups.
        ws.sheet_view.tabSelected = ws is active_ws
        _reset_sheet_view(ws)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    _normalize_generated_number_fonts(wb)
    out = io.BytesIO()
    wb.save(out)
    style_wb.close()
    return _sanitize_xlsx_package(out.getvalue())

def _fy_data_version(fy: str) -> tuple:
    """Cheap fingerprint of a financial year's rows, used to invalidate the cache."""
    conn = get_db_connection()
    try:
        row = conn.execute(
            "SELECT COUNT(*), COALESCE(MAX(last_modified), '') FROM amfi_metrics WHERE financial_year = ?",
            (fy,),
        ).fetchone()
    finally:
        conn.close()
    return (row[0], row[1])

def compile_dashboard_payload(fy: str) -> dict:
    """Read-only dashboard payload for GET endpoints.

    Compiling the workbook and re-parsing it into JSON is expensive (several
    seconds), so the result is memoized per financial year and only recomputed
    when that year's rows change. The payload shape is identical to
    ``dashboard_payload``; this never mutates the database.
    """
    version = _fy_data_version(fy)
    cached = _DASHBOARD_CACHE.get(fy)
    if cached and cached[0] == version:
        return cached[1]
    excel_bytes = compile_excel_for_fy(fy)
    payload = dashboard_payload(excel_bytes, fy=fy)
    payload["financialYear"] = fy
    with _DASHBOARD_CACHE_LOCK:
        _DASHBOARD_CACHE[fy] = (version, payload)
    return payload

def invalidate_dashboard_cache(fy: str | None = None) -> None:
    """Drop cached dashboard payloads; called after a successful upload."""
    with _DASHBOARD_CACHE_LOCK:
        if fy is None:
            _DASHBOARD_CACHE.clear()
        else:
            _DASHBOARD_CACHE.pop(fy, None)

def _compile_excel_for_fy_legacy(fy: str) -> bytes:
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM amfi_metrics WHERE financial_year = ?", (fy,))
    records = [dict(r) for r in cursor.fetchall()]

    if not records:
        try:
            start_year = int(fy.split("-")[0])
        except Exception:
            start_year = 2026
        prev_fy = f"{start_year - 1}-{start_year}"
        cursor.execute("""
            SELECT scheme_name, asset_type, scheme_structure, debt_equity, sales_prod_mis,
                   no_schemes, folios, funds_mobilized, redemption, net_inflow,
                   aum, avg_aum, seg_portfolios, seg_aum
            FROM amfi_metrics
            WHERE month = 3 AND year = ? AND financial_year = ?
        """, (start_year, prev_fy))
        march_records = cursor.fetchall()
        if march_records:
            baseline_records = [
                (
                    r["scheme_name"], r["asset_type"], r["scheme_structure"], r["debt_equity"], r["sales_prod_mis"],
                    r["no_schemes"], r["folios"], r["funds_mobilized"], r["redemption"], r["net_inflow"],
                    r["aum"], r["avg_aum"], r["seg_portfolios"], r["seg_aum"],
                    3, start_year, fy
                )
                for r in march_records
            ]
            cursor.executemany("""
                INSERT OR IGNORE INTO amfi_metrics (
                    scheme_name, asset_type, scheme_structure, debt_equity, sales_prod_mis,
                    no_schemes, folios, funds_mobilized, redemption, net_inflow,
                    aum, avg_aum, seg_portfolios, seg_aum,
                    month, year, financial_year
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, baseline_records)
            conn.commit()

            cursor.execute("SELECT * FROM amfi_metrics WHERE financial_year = ?", (fy,))
            records = [dict(r) for r in cursor.fetchall()]

    conn.close()

    if not records:
        raise ValueError(f"No records found in database for fiscal year: {fy}")

    try:
        start_year = int(fy.split("-")[0])
    except Exception:
        start_year = 2025

    records_by_seq = {}
    for r in records:
        seq = get_month_seq(r["month"], r["year"], start_year)
        records_by_seq.setdefault(seq, []).append(r)

    sorted_seqs = sorted(records_by_seq.keys())

    wb = load_workbook(TEMPLATE_PATH)
    ws_flat = wb[SHEET_FLAT] if SHEET_FLAT in wb.sheetnames else wb.worksheets[1]
    ws_form = wb[SHEET_FORM] if SHEET_FORM in wb.sheetnames else wb.worksheets[2]

    # capture template styles before clearing data columns
    flat_march_widths, flat_march_styles = extract_column_styles(ws_flat, 6, 2)
    flat_sep_widths, flat_sep_styles = extract_column_styles(ws_flat, 8, 1)
    flat_metric_widths, flat_metric_styles = extract_column_styles(ws_flat, 9, 9)

    cum_idx = 142 if ws_flat.max_column >= 143 else 19
    flat_cum_widths, flat_cum_styles = extract_column_styles(ws_flat, cum_idx, 2)

    grw_idx = 145 if ws_flat.max_column >= 147 else 22
    flat_growth_widths, flat_growth_styles = extract_column_styles(ws_flat, grw_idx, 3)

    form_march_widths, form_march_styles = extract_column_styles(ws_form, 3, 2)
    form_sep_widths, form_sep_styles = extract_column_styles(ws_form, 5, 1)
    form_metric_widths, form_metric_styles = extract_column_styles(ws_form, 6, 11)

    # capture template row data for attribute columns (cols 1-5)
    template_rows = {}
    for row in range(DATA_START_ROW, ws_flat.max_row + 1):
        name = ws_flat.cell(row, 1).value
        if not name or _is_summary_row(ws_flat, row):
            continue
        name_str = str(name).strip()
        if name_str.startswith("**") or name_str.startswith("##") or name_str == "-":
            continue
        template_rows[norm_key(name)] = {
            "scheme_name": name,
            "asset_type": ws_flat.cell(row, 2).value,
            "scheme_structure": ws_flat.cell(row, 3).value,
            "debt_equity": ws_flat.cell(row, 4).value,
            "sales_prod_mis": ws_flat.cell(row, 5).value,
        }

    # capture Grand Total row style tokens for macro headers
    total_row_idx = None
    for row in range(DATA_START_ROW, ws_flat.max_row + 1):
        if norm_key(ws_flat.cell(row, 1).value) == "grand total":
            total_row_idx = row
            break
    total_font = copy(ws_flat.cell(total_row_idx, 1).font) if total_row_idx else Font(bold=True, size=9, name="Calibri")
    total_fill = copy(ws_flat.cell(total_row_idx, 1).fill) if total_row_idx else PatternFill(fill_type="solid", fgColor="FFE699")
    total_nf = ws_flat.cell(total_row_idx, 6).number_format if total_row_idx else '_ * #,##0_ ;_ * \\-#,##0_ ;_ * "-"??_ ;_ @_ '

    # build scheme-to-category lookup and collect uncategorized schemes
    categorized_schemes = set()
    for _, schemes in MACRO_CATEGORIES:
        for s in schemes:
            categorized_schemes.add(norm_key(s))

    all_template_keys = list(template_rows.keys())
    uncategorized = [k for k in all_template_keys if k not in categorized_schemes]

    # clear flat sheet data area and rebuild sequentially
    ws_flat.delete_cols(6, ws_flat.max_column - 5)
    clean_merged_ranges(ws_flat, 6)

    # delete all data rows from the flat sheet (keep header rows 1-3)
    if ws_flat.max_row >= DATA_START_ROW:
        ws_flat.delete_rows(DATA_START_ROW, ws_flat.max_row - DATA_START_ROW + 1)

    # build month-sequenced DB maps
    seq_db_maps = {}
    for seq in sorted_seqs:
        group = records_by_seq[seq]
        seq_db_maps[seq] = {norm_key(r["scheme_name"]): r for r in group}

    # assemble data columns onto flat sheet
    flat_months_cols = []
    latest_month_key = ""
    prev_month_key = ""
    for seq in sorted_seqs:
        group = records_by_seq[seq]
        m_val = group[0]["month"]
        y_val = group[0]["year"]
        month_abbr = MONTH_ABBR[m_val]
        month_key = f"{month_abbr}'{str(y_val)[-2:]}"

        if seq == 0:
            latest_month_key = month_key
            col_start = ws_flat.max_column + 1
            for offset in range(2):
                apply_column_style(ws_flat, col_start + offset, flat_march_widths[offset], flat_march_styles[offset])
            ws_flat.cell(MONTH_ROW, col_start).value = month_key
            ws_flat.cell(HEADER_ROW, col_start).value = (
                "Net Assets Under Management as on "
                + datetime(y_val, m_val, 31).strftime("%B %d, %Y")
                + " (INR in crore)"
            )
            ws_flat.cell(HEADER_ROW, col_start + 1).value = (
                "Average Net Assets Under Management for the month "
                + datetime(y_val, m_val, 1).strftime("%B %Y")
                + " (INR in crore)"
            )
            ws_flat.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_start + 1)
            ws_flat.merge_cells(start_row=2, start_column=col_start, end_row=2, end_column=col_start + 1)
        else:
            prev_month_key = latest_month_key
            latest_month_key = month_key

            flat_sep = ws_flat.max_column + 1
            apply_column_style(ws_flat, flat_sep, flat_sep_widths[0], flat_sep_styles[0])
            ws_flat.cell(MONTH_ROW, flat_sep).value = "-"
            ws_flat.cell(HEADER_ROW, flat_sep).value = "-"

            flat_start = ws_flat.max_column + 1
            flat_months_cols.append(flat_start)
            for offset in range(9):
                apply_column_style(ws_flat, flat_start + offset, flat_metric_widths[offset], flat_metric_styles[offset])

            ws_flat.cell(MONTH_ROW, flat_start).value = month_key
            m_info = month_info(m_val, y_val)
            for offset, metric in enumerate(METRIC_ORDER):
                ws_flat.cell(HEADER_ROW, flat_start + offset).value = metric_label(metric, m_info)

            ws_flat.merge_cells(start_row=1, start_column=flat_start, end_row=1, end_column=flat_start + 8)
            ws_flat.merge_cells(start_row=2, start_column=flat_start, end_row=2, end_column=flat_start + 8)

    # add cumulative and growth columns
    if flat_months_cols:
        flat_sep = ws_flat.max_column + 1
        apply_column_style(ws_flat, flat_sep, flat_sep_widths[0], flat_sep_styles[0])
        ws_flat.cell(MONTH_ROW, flat_sep).value = "-"
        ws_flat.cell(HEADER_ROW, flat_sep).value = "-"

        cum_start = ws_flat.max_column + 1
        for offset in range(2):
            apply_column_style(ws_flat, cum_start + offset, flat_cum_widths[offset], flat_cum_styles[offset])

        ws_flat.cell(MONTH_ROW, cum_start).value = f"Apr'{str(start_year)[-2:]} to {latest_month_key}"
        ws_flat.cell(HEADER_ROW, cum_start).value = "Funds Mobilized  (INR in crore)"
        ws_flat.cell(HEADER_ROW, cum_start + 1).value = "Net Inflow (+ve)/Outflow (-ve)  (INR in crore)"
        ws_flat.merge_cells(start_row=1, start_column=cum_start, end_row=1, end_column=cum_start + 1)
        ws_flat.merge_cells(start_row=2, start_column=cum_start, end_row=2, end_column=cum_start + 1)

        flat_sep2 = ws_flat.max_column + 1
        apply_column_style(ws_flat, flat_sep2, flat_sep_widths[0], flat_sep_styles[0])
        ws_flat.cell(MONTH_ROW, flat_sep2).value = "-"
        ws_flat.cell(HEADER_ROW, flat_sep2).value = "-"

        grw_start = ws_flat.max_column + 1
        for offset in range(3):
            apply_column_style(ws_flat, grw_start + offset, flat_growth_widths[offset], flat_growth_styles[offset])

        ws_flat.cell(MONTH_ROW, grw_start).value = "-"
        ws_flat.cell(HEADER_ROW, grw_start).value = f"GS - Growth-({prev_month_key} Vs {latest_month_key})"
        ws_flat.cell(HEADER_ROW, grw_start + 1).value = f"NS - Growth-({prev_month_key} Vs {latest_month_key})"
        ws_flat.cell(HEADER_ROW, grw_start + 2).value = f"AUM - Growth-({prev_month_key} Vs {latest_month_key})"
    else:
        cum_start = None
        grw_start = None

    max_col = ws_flat.max_column

    # sequential row assembly with macro-category accordion groups
    row_cursor = DATA_START_ROW
    macro_header_rows = []

    def _write_scheme_row(ws, row_idx, scheme_key, db_maps, sorted_seqs, records_by_seq,
                          flat_months_cols, cum_start, grw_start, prev_idx_val, latest_idx_val):
        tpl = template_rows.get(scheme_key, {})
        ws.cell(row_idx, 1).value = tpl.get("scheme_name", scheme_key)
        ws.cell(row_idx, 2).value = tpl.get("asset_type", "")
        ws.cell(row_idx, 3).value = tpl.get("scheme_structure", "")
        ws.cell(row_idx, 4).value = tpl.get("debt_equity", "")
        ws.cell(row_idx, 5).value = tpl.get("sales_prod_mis", "")

        for seq in sorted_seqs:
            db_map = db_maps[seq]
            r_db = db_map.get(scheme_key)
            if seq == 0:
                col_start = 6
                if r_db:
                    ws.cell(row_idx, col_start).value = r_db["aum"]
                    ws.cell(row_idx, col_start + 1).value = r_db["avg_aum"]
            else:
                group = records_by_seq[seq]
                m_val = group[0]["month"]
                y_val = group[0]["year"]
                seq_offset = sorted_seqs.index(seq)
                month_col_idx = seq_offset - 1
                if month_col_idx < len(flat_months_cols):
                    flat_start = flat_months_cols[month_col_idx]
                    if r_db:
                        metrics_map = {
                            "no_schemes": r_db["no_schemes"],
                            "folios": r_db["folios"],
                            "funds_mobilized": r_db["funds_mobilized"],
                            "redemption": r_db["redemption"],
                            "net_inflow": r_db["net_inflow"],
                            "net_aum": r_db["aum"],
                            "avg_aum": r_db["avg_aum"],
                            "seg_portfolios": r_db["seg_portfolios"],
                            "seg_aum": r_db["seg_aum"],
                        }
                        for offset, metric in enumerate(METRIC_ORDER):
                            ws.cell(row_idx, flat_start + offset).value = metrics_map.get(metric)

        if flat_months_cols and cum_start:
            funds_refs = ",".join(f"{get_column_letter(c + 2)}{row_idx}" for c in flat_months_cols)
            net_refs = ",".join(f"{get_column_letter(c + 4)}{row_idx}" for c in flat_months_cols)
            ws.cell(row_idx, cum_start).value = f"=SUM({funds_refs})"
            ws.cell(row_idx, cum_start + 1).value = f"=SUM({net_refs})"

        if flat_months_cols and grw_start:
            if prev_idx_val == 6:
                ws.cell(row_idx, grw_start).value = "=0"
                ws.cell(row_idx, grw_start + 1).value = "=0"
            else:
                ws.cell(row_idx, grw_start).value = growth_formula(prev_idx_val + 2, latest_idx_val + 2, row_idx)
                ws.cell(row_idx, grw_start + 1).value = growth_formula(prev_idx_val + 4, latest_idx_val + 4, row_idx)
            prev_aum_col = 6 if prev_idx_val == 6 else prev_idx_val + 5
            ws.cell(row_idx, grw_start + 2).value = growth_formula(prev_aum_col, latest_idx_val + 5, row_idx)

    latest_idx = flat_months_cols[-1] if flat_months_cols else 6
    prev_idx = flat_months_cols[-2] if len(flat_months_cols) >= 2 else 6

    for cat_name, cat_schemes in MACRO_CATEGORIES:
        header_row = row_cursor
        ws_flat.cell(header_row, 1).value = cat_name
        for c in range(1, max_col + 1):
            cell = ws_flat.cell(header_row, c)
            cell.font = copy(total_font)
            cell.fill = copy(total_fill)
        for c in range(2, 6):
            ws_flat.cell(header_row, c).value = "-"
        macro_header_rows.append(header_row)
        row_cursor += 1

        child_rows = []
        for scheme_name in cat_schemes:
            sk = norm_key(scheme_name)
            if sk not in template_rows and sk not in {norm_key(r["scheme_name"]) for seq in sorted_seqs for r in records_by_seq[seq]}:
                continue
            _write_scheme_row(
                ws_flat, row_cursor, sk, seq_db_maps, sorted_seqs, records_by_seq,
                flat_months_cols, cum_start, grw_start, prev_idx, latest_idx,
            )
            child_rows.append(row_cursor)
            row_cursor += 1

        if child_rows:
            ws_flat.row_dimensions.group(child_rows[0], child_rows[-1], outline_level=1, hidden=True)

        # inject SUM formulas on macro header row
        if child_rows:
            first_child = child_rows[0]
            last_child = child_rows[-1]
            for c in range(6, max_col + 1):
                col_letter = get_column_letter(c)
                ws_flat.cell(header_row, c).value = f"=SUM({col_letter}{first_child}:{col_letter}{last_child})"
                ws_flat.cell(header_row, c).number_format = total_nf

    # write uncategorized scheme rows (no accordion)
    uncat_rows = []
    for sk in uncategorized:
        _write_scheme_row(
            ws_flat, row_cursor, sk, seq_db_maps, sorted_seqs, records_by_seq,
            flat_months_cols, cum_start, grw_start, prev_idx, latest_idx,
        )
        uncat_rows.append(row_cursor)
        row_cursor += 1

    # blank separator row
    row_cursor += 1

    # Grand Total row summing macro headers + uncategorized rows
    grand_total_row = row_cursor
    ws_flat.cell(grand_total_row, 1).value = "Grand Total"
    for c in range(1, max_col + 1):
        cell = ws_flat.cell(grand_total_row, c)
        cell.font = copy(total_font)
        cell.fill = copy(total_fill)
    for c in range(2, 6):
        ws_flat.cell(grand_total_row, c).value = "-"
    all_sum_rows = macro_header_rows + uncat_rows
    for c in range(6, max_col + 1):
        col_letter = get_column_letter(c)
        refs = "+".join(f"{col_letter}{hr}" for hr in all_sum_rows)
        ws_flat.cell(grand_total_row, c).value = f"={refs}"
        ws_flat.cell(grand_total_row, c).number_format = total_nf
    row_cursor += 1

    # outline grouping config
    ws_flat.sheet_properties.outlinePr = Outline(summaryBelow=False, showOutlineSymbols=True)

    # rebuild form sheet columns (unchanged logic)
    ws_form.delete_cols(3, ws_form.max_column - 2)
    clean_merged_ranges(ws_form, 3)

    for seq in sorted_seqs:
        group = records_by_seq[seq]
        m_val = group[0]["month"]
        y_val = group[0]["year"]
        month_abbr = MONTH_ABBR[m_val]
        month_key = f"{month_abbr}'{str(y_val)[-2:]}"
        db_map = seq_db_maps[seq]

        if seq == 0:
            col_start_form = ws_form.max_column + 1
            for offset in range(2):
                apply_column_style(ws_form, col_start_form + offset, form_march_widths[offset], form_march_styles[offset])
            ws_form.cell(MONTH_ROW, col_start_form).value = month_key
            ws_form.cell(HEADER_ROW, col_start_form).value = (
                "Net Assets Under Management as on "
                + datetime(y_val, m_val, 31).strftime("%B %d, %Y")
                + " (INR in crore)"
            )
            ws_form.cell(HEADER_ROW, col_start_form + 1).value = (
                "Average Net Assets Under Management for the month "
                + datetime(y_val, m_val, 1).strftime("%B %Y")
                + " (INR in crore)"
            )
            ws_form.merge_cells(start_row=2, start_column=col_start_form, end_row=2, end_column=col_start_form + 1)

            for row in range(DATA_START_ROW, ws_form.max_row + 1):
                if row in _FORM_SUBTOTAL_ROWS:
                    write_form_column_subtotals(ws_form, col_start_form)
                    write_form_column_subtotals(ws_form, col_start_form + 1)
                else:
                    s_name = ws_form.cell(row, 2).value
                    if s_name and not is_aggregate_scheme(s_name):
                        r_db = db_map.get(norm_key(s_name))
                        if r_db:
                            ws_form.cell(row, col_start_form).value = r_db["aum"]
                            ws_form.cell(row, col_start_form + 1).value = r_db["avg_aum"]
        else:
            form_sep = ws_form.max_column + 1
            apply_column_style(ws_form, form_sep, form_sep_widths[0], form_sep_styles[0])
            ws_form.cell(MONTH_ROW, form_sep).value = "-"
            ws_form.cell(HEADER_ROW, form_sep).value = "-"

            m_info = month_info(m_val, y_val)
            form_start = ws_form.max_column + 1
            for offset in range(11):
                apply_column_style(ws_form, form_start + offset, form_metric_widths[offset], form_metric_styles[offset])

            ws_form.cell(MONTH_ROW, form_start).value = f"Monthly Report for the month of {m_info['full_month']} {y_val} "
            ws_form.cell(HEADER_ROW, form_start).value = "Sr "
            ws_form.cell(HEADER_ROW, form_start + 1).value = "Scheme Name "
            for offset, metric in enumerate(METRIC_ORDER):
                ws_form.cell(HEADER_ROW, form_start + 2 + offset).value = metric_label(metric, m_info)

            ws_form.merge_cells(start_row=1, start_column=form_start, end_row=1, end_column=form_start + 10)
            ws_form.merge_cells(start_row=2, start_column=form_start, end_row=2, end_column=form_start + 10)

            for row in range(DATA_START_ROW, ws_form.max_row + 1):
                if row in _FORM_SUBTOTAL_ROWS:
                    for offset in range(9):
                        write_form_column_subtotals(ws_form, form_start + 2 + offset)
                else:
                    s_name = ws_form.cell(row, 2).value
                    if s_name and not is_aggregate_scheme(s_name):
                        r_db = db_map.get(norm_key(s_name))
                        ws_form.cell(row, form_start).value = ws_form.cell(row, 1).value
                        ws_form.cell(row, form_start + 1).value = ws_form.cell(row, 2).value
                        if r_db:
                            metrics_map = {
                                "no_schemes": r_db["no_schemes"],
                                "folios": r_db["folios"],
                                "funds_mobilized": r_db["funds_mobilized"],
                                "redemption": r_db["redemption"],
                                "net_inflow": r_db["net_inflow"],
                                "net_aum": r_db["aum"],
                                "avg_aum": r_db["avg_aum"],
                                "seg_portfolios": r_db["seg_portfolios"],
                                "seg_aum": r_db["seg_aum"],
                            }
                            for offset, metric in enumerate(METRIC_ORDER):
                                ws_form.cell(row, form_start + 2 + offset).value = metrics_map.get(metric)

    baseline_month_key = f"Mar'{str(start_year)[-2:]}"
    ws_flat.title = f"AMFI-{baseline_month_key} to {latest_month_key}"
    ws_form.title = f"AMFI-{baseline_month_key} to {latest_month_key}-AMFI form"
    wb.active = wb.worksheets.index(ws_flat)

    for ws in wb.worksheets:
        ws.sheet_view.topLeftCell = "A1"
        if ws.sheet_view.selection:
            ws.sheet_view.selection[0].activeCell = "A1"
            ws.sheet_view.selection[0].sqref = "A1"
        else:
            ws.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]
        pane = ws.sheet_view.pane
        if pane:
            x = int(pane.xSplit) if pane.xSplit else 0
            y = int(pane.ySplit) if pane.ySplit else 0
            pane.topLeftCell = f"{get_column_letter(x + 1)}{y + 1}"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

def rename_month_sheets(flat, form, latest_month: str) -> None:
    flat.title = f"{SHEET_FLAT_PREFIX}{latest_month}"
    form.title = f"{SHEET_FORM_PREFIX}{latest_month}{SHEET_FORM_SUFFIX}"

def growth_formula(previous_col: int | None, current_col: int | None, row: int) -> str:
    if not previous_col or not current_col or previous_col == current_col:
        return "=0"
    prev_col = get_column_letter(previous_col)
    curr_col = get_column_letter(current_col)
    return f"=IF({prev_col}{row}=0, 0, ({curr_col}{row}-{prev_col}{row})/{prev_col}{row})"

def build_summary_db(fy: str, wb) -> dict:
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        start_year = int(fy.split("-")[0])
    except Exception:
        start_year = 2025

    cursor.execute("""
        SELECT scheme_key, scheme_name, asset_type, scheme_structure, debt_equity, sales_prod_mis,
               month, year, funds_mobilized, redemption, net_inflow, aum, avg_aum
        FROM amfi_metrics
        WHERE financial_year = ?
    """, (fy,))

    raw_rows_by_key = {}
    for r in cursor.fetchall():
        if _is_reference_only_scheme(r["scheme_name"]):
            continue
        row = dict(r)
        row_key = _record_key(row)
        raw_rows_by_key[(row["month"], row["year"], row_key)] = row
    conn.close()
    grouped = {}
    for row in raw_rows_by_key.values():
        key = (row["month"], row["year"])
        item = grouped.setdefault(key, {
            "month": row["month"],
            "year": row["year"],
            "funds_mobilized": 0.0,
            "redemption": 0.0,
            "net_inflow": 0.0,
            "net_aum": 0.0,
            "avg_aum": 0.0,
        })
        item["funds_mobilized"] += row["funds_mobilized"] or 0.0
        item["redemption"] += row["redemption"] or 0.0
        item["net_inflow"] += row["net_inflow"] or 0.0
        item["net_aum"] += row["aum"] or 0.0
        item["avg_aum"] += row["avg_aum"] or 0.0

    rows = sorted(grouped.values(), key=lambda r: get_month_seq(r["month"], r["year"], start_year))

    series = []
    for r in rows:
        abbr = MONTH_ABBR[r["month"]]
        month_key = f"{abbr}'{str(r['year'])[-2:]}"
        series.append({
            "month": month_key,
            "funds_mobilized": round(r["funds_mobilized"] or 0.0, 2),
            "redemption": round(r["redemption"] or 0.0, 2),
            "net_inflow": round(r["net_inflow"] or 0.0, 2),
            "net_aum": round(r["net_aum"] or 0.0, 2),
            "avg_aum": round(r["avg_aum"] or 0.0, 2),
        })
        
    latest = series[-1] if series else {}
    return {
        "sheetCount": len(wb.sheetnames),
        "timeSeries": series,
        "latestMonth": latest.get("month"),
        "latestFundsMobilized": latest.get("funds_mobilized", 0),
        "latestNetInflow": latest.get("net_inflow", 0),
        "latestNetAum": latest.get("net_aum", 0),
    }

def _category_for_record(record: dict) -> str:
    matched = match_catalog_entry(
        record.get("scheme_name"),
        record.get("asset_type"),
        record.get("scheme_structure"),
        record.get("debt_equity"),
        record.get("sales_prod_mis"),
    )
    if matched:
        return matched["category"]
    return classify_macro_category(
        record.get("scheme_name"),
        record.get("asset_type"),
        record.get("scheme_structure"),
        record.get("debt_equity"),
        record.get("sales_prod_mis"),
        record.get("fund_type"),
    )

def _latest_previous_sequences(records: list[dict], start_year: int) -> tuple[int, int | None]:
    sequences = sorted({get_month_seq(row["month"], row["year"], start_year) for row in records})
    latest_seq = sequences[-1]
    previous_seq = sequences[-2] if len(sequences) > 1 else None
    return latest_seq, previous_seq

def build_category_and_scheme_summaries(fy: str) -> dict:
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT scheme_key, scheme_name, asset_type, scheme_structure, debt_equity, sales_prod_mis,
                   fund_type, month, year, funds_mobilized, redemption, net_inflow, aum, avg_aum
            FROM amfi_metrics
            WHERE financial_year = ?
        """, (fy,))
        records = [dict(row) for row in cursor.fetchall()]
    finally:
        conn.close()

    if not records:
        return {"categorySummary": [], "schemeSummary": []}

    start_year = int(fy.split("-")[0])
    latest_seq, previous_seq = _latest_previous_sequences(records, start_year)
    by_key = defaultdict(dict)
    for record in records:
        if _is_reference_only_scheme(record.get("scheme_name")):
            continue
        key = _record_key(record)
        seq = get_month_seq(record["month"], record["year"], start_year)
        by_key[key][seq] = record

    category_rows = {
        name: {
            "category": name,
            "schemeCount": 0,
            "latestAum": 0.0,
            "averageAum": 0.0,
            "monthlySales": 0.0,
            "monthlyRedemption": 0.0,
            "monthlyNetFlow": 0.0,
            "fytdSales": 0.0,
            "fytdRedemption": 0.0,
            "fytdNetFlow": 0.0,
            "aumGrowth": None,
        }
        for name in CATEGORY_NAMES
    }
    scheme_rows = []

    for scheme_key, seq_map in by_key.items():
        latest = seq_map.get(latest_seq)
        if not latest:
            continue
        previous = seq_map.get(previous_seq) if previous_seq is not None else None
        category = _category_for_record(latest)
        fund_type = latest.get("fund_type") or category
        latest_aum = float(latest.get("aum") or 0)
        previous_aum = float(previous.get("aum") or 0) if previous else 0
        aum_growth = None if previous_aum == 0 else (latest_aum - previous_aum) / previous_aum
        fytd_sales = _sum_values(record.get("funds_mobilized") for seq, record in seq_map.items() if seq > 0)
        fytd_redemption = _sum_values(record.get("redemption") for seq, record in seq_map.items() if seq > 0)
        fytd_net_flow = _sum_values(record.get("net_inflow") for seq, record in seq_map.items() if seq > 0)

        scheme_rows.append({
            "schemeKey": scheme_key,
            "schemeName": latest.get("scheme_name"),
            "category": category,
            "fundType": fund_type,
            "latestAum": round(latest_aum, 2),
            "averageAum": round(float(latest.get("avg_aum") or 0), 2),
            "monthlySales": round(float(latest.get("funds_mobilized") or 0), 2),
            "monthlyRedemption": round(float(latest.get("redemption") or 0), 2),
            "monthlyNetFlow": round(float(latest.get("net_inflow") or 0), 2),
            "fytdSales": round(fytd_sales, 2),
            "fytdRedemption": round(fytd_redemption, 2),
            "fytdNetFlow": round(fytd_net_flow, 2),
            "aumGrowth": None if aum_growth is None else round(aum_growth, 6),
        })

        bucket = category_rows.setdefault(category, {
            "category": category,
            "schemeCount": 0,
            "latestAum": 0.0,
            "averageAum": 0.0,
            "monthlySales": 0.0,
            "monthlyRedemption": 0.0,
            "monthlyNetFlow": 0.0,
            "fytdSales": 0.0,
            "fytdRedemption": 0.0,
            "fytdNetFlow": 0.0,
            "aumGrowth": None,
        })
        bucket["schemeCount"] += 1
        bucket["latestAum"] += latest_aum
        bucket["averageAum"] += float(latest.get("avg_aum") or 0)
        bucket["monthlySales"] += float(latest.get("funds_mobilized") or 0)
        bucket["monthlyRedemption"] += float(latest.get("redemption") or 0)
        bucket["monthlyNetFlow"] += float(latest.get("net_inflow") or 0)
        bucket["fytdSales"] += fytd_sales
        bucket["fytdRedemption"] += fytd_redemption
        bucket["fytdNetFlow"] += fytd_net_flow

    total_aum = sum(row["latestAum"] for row in category_rows.values())
    for category, bucket in category_rows.items():
        latest_members = [
            seq_map.get(latest_seq)
            for scheme_key, seq_map in by_key.items()
            if seq_map.get(latest_seq) and _category_for_record(seq_map[latest_seq]) == category
        ]
        previous_total = 0.0
        if previous_seq is not None:
            for scheme_key, seq_map in by_key.items():
                latest = seq_map.get(latest_seq)
                previous = seq_map.get(previous_seq)
                if latest and previous and _category_for_record(latest) == category:
                    previous_total += float(previous.get("aum") or 0)
        bucket["aumShare"] = 0 if total_aum == 0 else bucket["latestAum"] / total_aum
        bucket["aumGrowth"] = None if previous_total == 0 else (bucket["latestAum"] - previous_total) / previous_total
        for key in ("latestAum", "averageAum", "monthlySales", "monthlyRedemption", "monthlyNetFlow", "fytdSales", "fytdRedemption", "fytdNetFlow"):
            bucket[key] = round(bucket[key], 2)
        bucket["aumShare"] = round(bucket["aumShare"], 6)
        bucket["aumGrowth"] = None if bucket["aumGrowth"] is None else round(bucket["aumGrowth"], 6)

    ordered_categories = [category_rows[name] for name in CATEGORY_NAMES if name in category_rows]
    scheme_rows.sort(key=lambda row: row["latestAum"], reverse=True)
    return {
        "categorySummary": ordered_categories,
        "schemeSummary": scheme_rows,
    }

def build_ns_analysis_summary(fy: str | None) -> dict:
    if not fy:
        return {
            "previousMonth": "-",
            "currentMonth": "-",
            "total": _ns_row("TOTAL", 0, 0, 3),
            "headlineRows": [],
            "equityRows": [],
            "hybridRows": [],
        }
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM amfi_metrics WHERE financial_year = ?", (fy,))
        records = [dict(row) for row in cursor.fetchall()]
    finally:
        conn.close()

    if not records:
        return {
            "previousMonth": "-",
            "currentMonth": "-",
            "total": _ns_row("TOTAL", 0, 0, 3),
            "headlineRows": [],
            "equityRows": [],
            "hybridRows": [],
        }
    start_year = int(fy.split("-")[0])
    records_by_seq = defaultdict(list)
    for record in records:
        records_by_seq[get_month_seq(record["month"], record["year"], start_year)].append(record)
    return build_ns_analysis_from_records(records_by_seq, sorted(records_by_seq))

def build_sip_summary(workbook_bytes: bytes) -> dict:
    wb = load_workbook(io.BytesIO(workbook_bytes), data_only=True)
    try:
        ws = wb[SHEET_SIP] if SHEET_SIP in wb.sheetnames else wb.worksheets[0]
        monthly = []
        sip_stats_rows = []
        for row in range(22, ws.max_row + 1):
            month_value = ws.cell(row, 1).value
            new_registrations = number_or_none(ws.cell(row, 3).value)
            discontinued = number_or_none(ws.cell(row, 4).value)
            contributing_accounts = number_or_none(ws.cell(row, 5).value)
            if (
                month_value not in (None, "")
                and any(value is not None for value in (new_registrations, discontinued, contributing_accounts))
            ):
                label = (
                    f"{MONTH_ABBR[month_value.month]}'{str(month_value.year)[-2:]}"
                    if isinstance(month_value, datetime)
                    else str(month_value)
                )
                sip_stats_rows.append({
                    "label": label,
                    "newRegistrations": new_registrations,
                    "discontinued": discontinued,
                    "contributingAccounts": contributing_accounts,
                    "isSummary": not isinstance(month_value, datetime),
                })
            if not isinstance(month_value, datetime):
                continue
            monthly.append({
                "month": f"{MONTH_ABBR[month_value.month]}'{str(month_value.year)[-2:]}",
                "date": month_value.isoformat(),
                "outstandingAccounts": number_or_none(ws.cell(row, 2).value),
                "newRegistrations": number_or_none(ws.cell(row, 3).value),
                "discontinued": number_or_none(ws.cell(row, 4).value),
                "contributingAccounts": number_or_none(ws.cell(row, 5).value),
                "contribution": number_or_none(ws.cell(row, 6).value),
                "aum": number_or_none(ws.cell(row, 7).value),
            })
        monthly.sort(key=lambda row: row["date"])

        annual = []
        for col in range(2, ws.max_column + 1):
            fy_label = ws.cell(4, col).value
            contribution = number_or_none(ws.cell(5, col).value)
            if contribution is None:
                contribution = _sum_values(ws.cell(row, col).value for row in range(6, 18))
            if fy_label and contribution is not None:
                annual.append({"financialYear": str(fy_label), "contribution": contribution})

        contribution_rows = [row for row in monthly if row.get("contribution") is not None]
        aum_rows = [row for row in monthly if row.get("aum") is not None]
        account_rows = [row for row in monthly if row.get("outstandingAccounts") is not None]
        new_registration_rows = [row for row in monthly if row.get("newRegistrations") is not None]
        discontinued_rows = [row for row in monthly if row.get("discontinued") is not None]
        contributing_rows = [row for row in monthly if row.get("contributingAccounts") is not None]
        latest = contribution_rows[-1] if contribution_rows else (monthly[-1] if monthly else {})
        previous = contribution_rows[-2] if len(contribution_rows) > 1 else {}
        contribution_growth = None
        if previous.get("contribution"):
            contribution_growth = (latest.get("contribution", 0) - previous["contribution"]) / previous["contribution"]
        return {
            "latestMonth": latest.get("month"),
            "latestContribution": round(latest.get("contribution", 0), 2),
            "latestAum": round(aum_rows[-1]["aum"], 2) if aum_rows else None,
            "latestAumMonth": aum_rows[-1]["month"] if aum_rows else None,
            "latestOutstandingAccounts": round(account_rows[-1]["outstandingAccounts"], 2) if account_rows else None,
            "latestOutstandingAccountsMonth": account_rows[-1]["month"] if account_rows else None,
            "latestNewRegistrations": round(new_registration_rows[-1]["newRegistrations"], 2) if new_registration_rows else None,
            "latestDiscontinued": round(discontinued_rows[-1]["discontinued"], 2) if discontinued_rows else None,
            "latestContributingAccounts": round(contributing_rows[-1]["contributingAccounts"], 2) if contributing_rows else None,
            "latestSipStatsMonth": (
                contributing_rows[-1]["month"]
                if contributing_rows
                else latest.get("month")
            ),
            "contributionGrowth": None if contribution_growth is None else round(contribution_growth, 6),
            "sipStatsRows": sip_stats_rows,
            "monthlySeries": monthly,
            "annualContributions": annual,
        }
    finally:
        wb.close()

def build_dashboard_insights(workbook_bytes: bytes, fy: str | None, wb) -> dict:
    summary = build_summary(wb, fy=fy)
    insights = {
        "summary": summary,
        "timeSeries": summary.get("timeSeries", []),
        "categorySummary": [],
        "schemeSummary": [],
        "sipSummary": build_sip_summary(workbook_bytes),
        "nsAnalysis": build_ns_analysis_summary(fy),
    }
    if fy:
        insights.update(build_category_and_scheme_summaries(fy))
    return insights

def dashboard_payload(workbook_bytes: bytes, warnings: list[str] | None = None, upload_month: str | None = None, fy: str | None = None) -> dict:
    wb = load_workbook(io.BytesIO(workbook_bytes), data_only=False)
    try:
        sheets = sheet_payloads(workbook_bytes, wb.sheetnames)
        insights = build_dashboard_insights(workbook_bytes, fy, wb)
        return {
            "sheets": sheets,
            **insights,
            "warnings": warnings or [],
            "uploadMonth": upload_month,
            "generatedAt": datetime.utcnow().isoformat(timespec="seconds") + "Z",
        }
    finally:
        wb.close()

def sheet_payloads(workbook_bytes: bytes, sheet_names: list[str]) -> dict:
    metadata_wb = load_workbook(io.BytesIO(workbook_bytes), data_only=True)
    sheets = {}
    for sheet_name in sheet_names:
        df = pd.read_excel(io.BytesIO(workbook_bytes), sheet_name=sheet_name, header=None).fillna("")
        rows = df.values.tolist()
        max_col = len(rows[0]) if rows else 0
        ws = metadata_wb[sheet_name]
        sheets[sheet_name] = {
            "name": sheet_name,
            "maxRow": len(rows),
            "maxColumn": max_col,
            "columns": rows[HEADER_ROW - 1] if len(rows) >= HEADER_ROW else [get_column_letter(c) for c in range(1, max_col + 1)],
            "rows": rows,
            "rowMeta": [
                {
                    "outlineLevel": ws.row_dimensions[index + 1].outlineLevel,
                    "hidden": bool(ws.row_dimensions[index + 1].hidden),
                    "isCategory": str(ws.cell(index + 1, 1).value or "").strip() in CATEGORY_NAMES,
                }
                for index in range(len(rows))
            ],
        }
    metadata_wb.close()
    return sheets

def _source_upload_key(row: dict) -> str:
    if _is_fof_domestic_scheme(row.get("scheme")):
        return "__fof_domestic__"
    return row.get("scheme_key") or stable_scheme_key(
        row.get("scheme"),
        row.get("asset_type"),
        row.get("scheme_structure"),
    )

def _flat_row_keys(ws) -> dict[str, int]:
    keys = {}
    for row in range(DATA_START_ROW, ws.max_row + 1):
        name = clean_scheme(ws.cell(row, 1).value)
        if not name:
            continue
        if _is_fof_domestic_scheme(name):
            keys["__fof_domestic__"] = row
            continue
        if is_aggregate_scheme(name) or name in CATEGORY_NAMES:
            continue
        key = stable_scheme_key(
            name,
            cell_value(ws, row, 2),
            cell_value(ws, row, 3),
        )
        keys.setdefault(key, row)
    return keys

def _current_flat_metric_columns(ws, month: int, year: int) -> dict[str, int]:
    month_key = f"{MONTH_ABBR[month]}'{str(year)[-2:]}"
    for block in month_blocks(ws):
        if block["month"] == month_key:
            return metric_columns(ws, block["start"], block["end"])
    return {}

def _current_form_metric_columns(ws, month: int, year: int) -> dict[str, int]:
    target = f"monthly report for the month of {calendar.month_name[month].lower()} {year}"
    for col in range(1, ws.max_column + 1):
        if target in norm(ws.cell(MONTH_ROW, col).value):
            return metric_columns(ws, col + 2, col + 2 + len(METRIC_ORDER) - 1)
    return {}

def _close_enough(expected, actual, tolerance: float = 0.02) -> bool:
    expected_value = number_or_none(expected)
    actual_value = number_or_none(actual)
    if expected_value is None and actual_value is None:
        return True
    if expected_value is None or actual_value is None:
        return False
    return abs(float(expected_value) - float(actual_value)) <= tolerance

def _compare_metric_cells(
    errors: list[str],
    label: str,
    expected_metrics: dict,
    ws,
    row: int,
    columns: dict[str, int],
    displays: dict | None = None,
) -> None:
    displays = displays or {}
    for metric in METRIC_ORDER:
        if metric not in expected_metrics or metric not in columns:
            continue
        expected = expected_metrics[metric]
        actual = ws.cell(row, columns[metric]).value
        if metric in displays:
            expected_display = str(displays[metric]).strip()
            if str(actual).strip() != expected_display:
                errors.append(f"{label} {metric}: expected display {expected_display}, got {actual}")
                continue
        if not _close_enough(expected, actual):
            errors.append(f"{label} {metric}: expected {expected}, got {actual}")

def validate_uploaded_month_against_generated(
    fy: str,
    month: int,
    year: int,
    source_rows: list[dict],
    conn: sqlite3.Connection | None = None,
) -> list[str]:
    if not source_rows:
        return []
    workbook_bytes = compile_excel_for_fy(fy, conn=conn)
    wb = load_workbook(io.BytesIO(workbook_bytes), data_only=False)
    errors: list[str] = []
    try:
        flat = flat_sheet(wb)
        form = form_sheet(wb)
        flat_rows = _flat_row_keys(flat)
        form_rows = {key: row for row, key in _form_row_keys(form).items()}
        flat_columns = _current_flat_metric_columns(flat, month, year)
        form_columns = _current_form_metric_columns(form, month, year)
        for source in source_rows:
            key = _source_upload_key(source)
            label = source.get("scheme") or key
            displays = source.get("metric_displays") or {}
            if key in flat_rows:
                _compare_metric_cells(
                    errors,
                    f"Main sheet {label}",
                    source.get("metrics", {}),
                    flat,
                    flat_rows[key],
                    flat_columns,
                    displays if key == "__fof_domestic__" else None,
                )
            if key not in form_rows:
                errors.append(f"AMFI form missing row for {label}")
                continue
            _compare_metric_cells(
                errors,
                f"AMFI form {label}",
                source.get("metrics", {}),
                form,
                form_rows[key],
                form_columns,
                displays if key == "__fof_domestic__" else None,
            )
    finally:
        wb.close()
    return errors

_SKIP_LABELS = ("grand total", "total", "fund of funds", "growth")

def _is_summary_row(ws, row: int) -> bool:
    value = norm_key(ws.cell(row, 1).value)
    if not value:
        return True
    return any(label in value for label in _SKIP_LABELS)

def build_summary(wb, fy: str | None = None) -> dict:
    if fy:
        try:
            return build_summary_db(fy, wb)
        except Exception:
            LOGGER.warning("Falling back to workbook-derived summary for %s", fy, exc_info=True)

    flat = flat_sheet(wb)
    blocks = month_blocks(flat)
    metric_keys = ("funds_mobilized", "redemption", "net_inflow", "net_aum", "avg_aum")
    series = []
    for block in blocks:
        cols = metric_columns(flat, block["start"], block["end"])
        sums = {k: 0.0 for k in metric_keys}
        for row in range(DATA_START_ROW, flat.max_row + 1):
            if _is_summary_row(flat, row):
                continue
            for key in metric_keys:
                col = cols.get(key)
                if col:
                    val = flat.cell(row, col).value
                    if isinstance(val, (int, float)):
                        sums[key] += float(val)
        sums = {k: round(v, 2) for k, v in sums.items()}
        series.append({"month": block["month"], **sums})

    latest = series[-1] if series else {}
    return {
        "sheetCount": len(wb.sheetnames),
        "timeSeries": series,
        "latestMonth": latest.get("month"),
        "latestFundsMobilized": latest.get("funds_mobilized", 0),
        "latestNetInflow": latest.get("net_inflow", 0),
        "latestNetAum": latest.get("net_aum", 0),
    }

def flat_sheet(wb):
    return sheet_by_name(wb, SHEET_FLAT, "AMFI-Mar'", exclude_suffix=SHEET_FORM_SUFFIX)

def form_sheet(wb):
    return sheet_by_name(wb, SHEET_FORM, "AMFI-Mar'", suffix=SHEET_FORM_SUFFIX)

def sheet_by_name(wb, exact: str, prefix: str, suffix: str | None = None, exclude_suffix: str | None = None):
    if exact in wb.sheetnames:
        return wb[exact]
    for name in wb.sheetnames:
        if not name.startswith(prefix):
            continue
        if suffix and not name.endswith(suffix):
            continue
        if exclude_suffix and name.endswith(exclude_suffix):
            continue
        return wb[name]
    raise KeyError(exact)

def parse_upload(upload_bytes: bytes, filename: str) -> tuple[list[dict], dict, list[str]]:
    wb = read_upload_workbook(upload_bytes)
    candidates = []
    for ws in wb.worksheets:
        header_row, columns = detect_header(ws)
        if header_row and "scheme" in columns and len([k for k in columns if k in METRIC_ORDER]) >= 3:
            candidates.append((ws, header_row, columns))
    if not candidates:
        raise ValueError("No upload sheet contained a recognizable AMFI metric header row.")

    ws, header_row, columns = max(candidates, key=lambda item: item[0].max_row * len(item[2]))
    header_text = " ".join(str(ws.cell(header_row, c).value or "") for c in range(1, ws.max_column + 1))
    month_info = infer_month(filename, header_text)
    rows = extract_scheme_rows(ws, header_row, columns)
    return rows, month_info, []

def read_upload_workbook(upload_bytes: bytes):
    file_bytes = io.BytesIO(upload_bytes)
    try:
        frames = pd.read_excel(file_bytes, sheet_name=None, header=None, engine="openpyxl")
    except Exception:
        file_bytes.seek(0)
        try:
            frames = pd.read_excel(file_bytes, sheet_name=None, header=None, engine="xlrd")
        except Exception:
            if not _env_bool("AMFI_ALLOW_HTML_UPLOAD", default=False):
                raise ValueError("Uploaded file is not a readable Excel workbook.")
            file_bytes.seek(0)
            frames = {"Upload": pd.read_html(file_bytes)[0]}
    return frames_to_workbook(frames)

def frames_to_workbook(frames: dict[str, pd.DataFrame]):
    wb = Workbook()
    wb.remove(wb.active)
    for name, df in frames.items():
        ws = wb.create_sheet(str(name)[:31])
        for row in df.fillna("").values.tolist():
            ws.append(row)
    return wb

def detect_header(ws) -> tuple[int | None, dict]:
    for row_idx in range(1, min(ws.max_row, 50) + 1):
        columns = {}
        for col_idx in range(1, ws.max_column + 1):
            label = norm(ws.cell(row_idx, col_idx).value)
            key = classify_header(label)
            if key and key not in columns:
                columns[key] = col_idx
        if "scheme" in columns:
            return row_idx, columns
    return None, {}

def classify_header(label: str) -> str | None:
    if not label:
        return None
    if "scheme name" in label:
        return "scheme"
    if label == "asset type" or label.endswith("asset type"):
        return "asset_type"
    if "open ended" in label or "closed ended" in label:
        return "scheme_structure"
    if "debt / equity" in label or "debt equity" in label:
        return "debt_equity"
    if "sales_prod_mis" in label or "sales prod mis" in label:
        return "sales_prod_mis"
    if label == "fund type":
        return "fund_type"
    if "no. of schemes" in label or "number of schemes" in label:
        return "no_schemes"
    if "no. of folios" in label or "number of folios" in label:
        return "folios"
    if "funds mobilized" in label:
        return "funds_mobilized"
    if "repurchase" in label or "redemption" in label:
        return "redemption"
    if "net inflow" in label or "outflow" in label:
        return "net_inflow"
    if "average net assets" in label or "average aum" in label:
        return "avg_aum"
    if "segregated portfolios created" in label:
        return "seg_portfolios"
    if "segregated portfolio" in label and "net assets" in label:
        return "seg_aum"
    if "net assets under management" in label or "net aum" in label:
        return "net_aum"
    return None

def extract_scheme_rows(ws, header_row: int, columns: dict) -> list[dict]:
    rows = []
    parent = None
    current_structure = None
    current_asset_type = None
    scheme_col = columns["scheme"]
    for row_idx in range(header_row + 1, ws.max_row + 1):
        scheme = clean_scheme(ws.cell(row_idx, scheme_col).value)
        if not scheme:
            continue
        raw_serial = str(ws.cell(row_idx, max(1, scheme_col - 1)).value or "").strip()
        serial = norm_key(ws.cell(row_idx, max(1, scheme_col - 1)).value)
        scheme_key = norm_key(scheme)
        metrics = {key: number_or_none(ws.cell(row_idx, columns[key]).value) for key in METRIC_ORDER if key in columns}
        metric_displays = {
            key: str(ws.cell(row_idx, columns[key]).value).strip()
            for key in METRIC_ORDER
            if key in columns and str(ws.cell(row_idx, columns[key]).value or "").strip().startswith("##")
        }
        has_metrics = any(value is not None for value in metrics.values())
        if serial in {"a", "b", "c"} and (
            "open ended" in scheme_key
            or "close ended" in scheme_key
            or "interval" in scheme_key
        ):
            current_structure = scheme
            current_asset_type = None
            parent = scheme
            continue
        if not has_metrics:
            if serial:
                current_asset_type = scheme
            parent = scheme
            continue
        if is_aggregate_scheme(scheme):
            continue
        asset_type = cell_value(ws, row_idx, columns.get("asset_type")) or current_asset_type or parent
        scheme_structure = cell_value(ws, row_idx, columns.get("scheme_structure")) or current_structure
        if raw_serial in {"I", "II", "III", "IV", "V"} and has_metrics:
            if norm_key(current_structure) == "interval schemes":
                asset_type = current_structure
                scheme_structure = current_structure
            elif norm_key(current_structure) == "close ended schemes" and scheme_key == "other schemes":
                asset_type = scheme
                scheme_structure = current_structure
        record = {
            "row": row_idx,
            "scheme": scheme,
            "parent": parent,
            "asset_type": asset_type,
            "scheme_structure": scheme_structure,
            "debt_equity": cell_value(ws, row_idx, columns.get("debt_equity")),
            "sales_prod_mis": cell_value(ws, row_idx, columns.get("sales_prod_mis")),
            "fund_type": cell_value(ws, row_idx, columns.get("fund_type")),
            "metrics": metrics,
            "metric_displays": metric_displays,
        }
        rows.append(enrich_scheme_record(record))
    return rows

def month_blocks(ws) -> list[dict]:
    blocks = []
    for col in range(1, ws.max_column + 1):
        value = ws.cell(MONTH_ROW, col).value
        if is_month_key(value) and has_metric_header(ws, col):
            end = col
            while end + 1 <= ws.max_column and ws.cell(HEADER_ROW, end + 1).value != "-":
                end += 1
            sep = col - 1 if col > 1 and ws.cell(HEADER_ROW, col - 1).value == "-" else None
            metric_count = len(metric_columns(ws, col, min(end, col + len(METRIC_ORDER) - 1)))
            limit = 2 if metric_count <= 2 else len(METRIC_ORDER)
            blocks.append({"month": value, "start": col, "end": min(end, col + limit - 1), "separator": sep})
    return blocks

def form_blocks(ws) -> list[dict]:
    blocks = []
    for col in range(1, ws.max_column + 1):
        value = str(ws.cell(MONTH_ROW, col).value or "")
        if value.lower().startswith("monthly report"):
            sep = col - 1 if col > 1 and ws.cell(HEADER_ROW, col - 1).value == "-" else None
            blocks.append({"start": col, "end": col + len(METRIC_ORDER) + 1, "separator": sep})
    return blocks

def has_metric_header(ws, col: int) -> bool:
    window = " ".join(norm(ws.cell(HEADER_ROW, c).value) for c in range(col, min(ws.max_column, col + 9) + 1))
    return "funds mobilized" in window or "no. of schemes" in window or "net assets under" in window

def metric_columns(ws, start_col: int, end_col: int) -> dict:
    columns = {}
    for col in range(start_col, min(ws.max_column, end_col) + 1):
        key = classify_header(norm(ws.cell(HEADER_ROW, col).value))
        if key in METRIC_ORDER and key not in columns:
            columns[key] = col
    return columns

def infer_month(filename: str, text: str) -> dict:
    patterns = [
        r"month\s+of\s+([A-Za-z]+)\s+(20\d{2})",
        r"as\s+on\s+([A-Za-z]+)\s+\d{1,2},?\s+(20\d{2})",
        r"([A-Za-z]{3,9})[' -]*(\d{2,4})",
    ]
    for source in (filename, f"{filename} {text}"):
        for pattern in patterns:
            match = re.search(pattern, source, flags=re.IGNORECASE)
            if not match:
                continue
            month_raw, year_raw = match.group(1).lower(), match.group(2)
            month_num = MONTHS.get(month_raw[:3]) or MONTHS.get(month_raw)
            if month_num:
                year = int(year_raw) if len(year_raw) == 4 else 2000 + int(year_raw)
                return month_info(month_num, year)
    raise ValueError("Unable to determine report month and year from the uploaded filename or workbook header.")

def month_info(month_num: int, year: int) -> dict:
    abbr = MONTH_ABBR[month_num]
    last_day = calendar.monthrange(year, month_num)[1]
    full_month = datetime(year, month_num, 1).strftime("%B")
    return {
        "month": month_num,
        "year": year,
        "key": f"{abbr}'{str(year)[-2:]}",
        "full_month": full_month,
        "date": f"{full_month} {last_day}, {year}",
    }

def metric_label(metric: str, month_info: dict) -> str:
    return METRIC_LABELS[metric].format(date=month_info["date"], month=month_info["full_month"], year=month_info["year"])

def is_month_key(value) -> bool:
    return isinstance(value, str) and bool(re.fullmatch(r"[A-Z][a-z]{2}'\d{2}", value.strip()))

def norm(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower().replace("/", " / "))

def norm_key(value) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").strip().lower()).strip()

def clean_scheme(value) -> str:
    text = str(value or "").strip()
    return re.sub(r"\s+", " ", text)

def is_aggregate_scheme(value) -> bool:
    key = norm_key(value)
    return key.startswith("sub total") or key.startswith("subtotal") or key.startswith("total ") or key == "grand total"

def number_or_none(value):
    if value is None or value == "-":
        return None
    if pd.isna(value):
        return None
    if isinstance(value, (int, float)):
        return value
    cleaned = re.sub(r"^[#\s]+", "", str(value).replace(",", "").strip())
    try:
        return float(cleaned)
    except ValueError:
        return None

def cell_value(ws, row: int, col: int | None):
    if not col:
        return None
    return ws.cell(row, col).value
